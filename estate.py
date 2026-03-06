import streamlit as st
import pandas as pd
import mysql.connector
from mysql.connector import Error
from datetime import datetime, date
from fpdf import FPDF
import io, os, re, traceback, smtplib, time, subprocess, tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Sunshine Estate", page_icon="🏠", layout="wide")

# Service charge master config
SERVICE_CHARGES = {
    'SVC':   {'name': 'Service Charge',   'amount': 240000.00, 'freq': 'Annual'},
    'DEV':   {'name': 'Development Levy', 'amount': 50000.00,  'freq': 'Annual'},
    'PARTY': {'name': 'Year End Party',   'amount': 0.00,      'freq': 'Annual'},
    'INFRA': {'name': 'Infrastructure',   'amount': 2000000.00,'freq': 'One Off'},
    'LEGAL': {'name': 'Legal Fee',        'amount': 50000.00,  'freq': 'One Off'},
    'TRANSF':{'name': 'Transformer',      'amount': 50000.00,  'freq': 'One Off'},
    'LIGHT': {'name': 'Light Connection', 'amount': 100000.00, 'freq': 'One Off'},
}

# ── Helpers ──────────────────────────────────────────────────────
def validate_email(email):
    if not email: return True
    return bool(re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email))

def validate_phone(phone):
    if not phone: return False
    clean = re.sub(r'[^\d+]', '', phone)
    return bool(re.match(r'^(\+?234|0)?[7-9][0-1]\d{8}$', clean))

def format_currency(amount):
    try: return f"₦{float(amount):,.2f}"
    except: return "₦0.00"

def format_currency_pdf(amount):
    try: return f"N{float(amount):,.2f}"
    except: return "N0.00"

def safe_str(val, default=''):
    if val is None or (isinstance(val, float) and pd.isna(val)): return default
    return str(val).strip()

def safe_float(val, default=0.0):
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)): return default
        return float(val)
    except: return default

def safe_date(val):
    if val is None or (isinstance(val, float) and pd.isna(val)): return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    try: return pd.to_datetime(val).date()
    except: return None

# ── Database ─────────────────────────────────────────────────────
def get_db():
    try:
        return mysql.connector.connect(
            host="nozomi.proxy.rlwy.net", user="root",
            password="oaHtEqQJQpMhKhKArkXWXKiyVkeTRKJL",
            database="sunshine_estate", port=41745,
            autocommit=False, ssl_verify_cert=False, connection_timeout=30
        )
    except Error as e:
        st.error(f"Database Error: {e}")
        return None

# ════════════════════════════════════════════════════════════════
# ESTATE MANAGER
# ════════════════════════════════════════════════════════════════
class EstateManager:
    def __init__(self):
        self.conn = get_db()

    def table_exists(self, t):
        cur = self.conn.cursor()
        try:
            cur.execute(f"SHOW TABLES LIKE '{t}'")
            return cur.fetchone() is not None
        finally: cur.close()

    # ── Stats ────────────────────────────────────────────────────
    def get_stats(self):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT COUNT(*) as v FROM properties"); props = cur.fetchone()['v']
            cur.execute("SELECT COUNT(*) as v FROM residents WHERE is_active=1"); residents = cur.fetchone()['v']
            cur.execute("""SELECT COALESCE(SUM(amount_paid),0) as v FROM payments
                           WHERE MONTH(payment_date)=MONTH(CURDATE()) AND YEAR(payment_date)=YEAR(CURDATE())""")
            monthly = cur.fetchone()['v']
            cur.execute("SELECT COALESCE(SUM(balance),0) as v FROM balances WHERE year=YEAR(CURDATE())")
            outstanding = cur.fetchone()['v']
            cur.execute("""SELECT p.*, r.name, pt.code as payment_code, pt.name as payment_type,
                                  pr.house_no, s.name as street
                           FROM payments p JOIN residents r ON p.resident_id=r.id
                           LEFT JOIN payment_types pt ON p.payment_type_id=pt.id
                           LEFT JOIN properties pr ON p.property_id=pr.id
                           LEFT JOIN streets s ON pr.street_id=s.id
                           ORDER BY p.id DESC LIMIT 10""")
            recent = cur.fetchall()
            return dict(props=props, residents=residents, monthly=monthly, outstanding=outstanding, recent=recent)
        except: return dict(props=0, residents=0, monthly=0, outstanding=0, recent=[])
        finally: cur.close()

    # ── Properties ───────────────────────────────────────────────
    def get_properties(self):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("""SELECT p.*, s.name as street,
                                  pt.name as type_name, pt.name as type,
                                  GROUP_CONCAT(r.name SEPARATOR ', ') as resident_names,
                                  GROUP_CONCAT(r.name SEPARATOR ', ') as residents
                           FROM properties p
                           JOIN streets s ON p.street_id=s.id
                           JOIN property_types pt ON p.type_id=pt.id
                           LEFT JOIN residents r ON r.property_id=p.id AND r.is_active=1
                           GROUP BY p.id ORDER BY s.name, p.house_no""")
            return cur.fetchall()
        finally: cur.close()

    def get_property_deletion_info(self, property_id):
        cur = self.conn.cursor(dictionary=True)
        try:
            info = {'can_delete': True, 'active_residents': 0,
                    'inactive_residents': 0, 'total_payments': 0, 'reasons': []}
            cur.execute("SELECT COUNT(*) as c FROM residents WHERE property_id=%s AND is_active=1", (property_id,))
            info['active_residents'] = cur.fetchone()['c']
            cur.execute("SELECT COUNT(*) as c FROM residents WHERE property_id=%s AND is_active=0", (property_id,))
            info['inactive_residents'] = cur.fetchone()['c']
            cur.execute("SELECT COUNT(*) as c FROM payments WHERE property_id=%s", (property_id,))
            info['total_payments'] = cur.fetchone()['c']
            if info['active_residents'] > 0:
                info['can_delete'] = False
                info['reasons'].append(f"{info['active_residents']} active resident(s)")
            if info['inactive_residents'] > 0:
                info['can_delete'] = False
                info['reasons'].append(f"{info['inactive_residents']} inactive resident(s)")
            if info['total_payments'] > 0:
                info['can_delete'] = False
                info['reasons'].append(f"{info['total_payments']} payment record(s)")
            return info
        except: return None
        finally: cur.close()

    def add_property(self, house_no, street_id, type_id, status):
        cur = self.conn.cursor()
        try:
            cur.execute("INSERT INTO properties (house_no,street_id,type_id,status) VALUES(%s,%s,%s,%s)",
                        (house_no, street_id, type_id, status))
            self.conn.commit(); return True, "✅ Property added successfully"
        except Error as e:
            self.conn.rollback(); return False, f"Error: {e}"
        finally: cur.close()

    def update_property(self, pid, house_no, street_id, type_id, status):
        cur = self.conn.cursor()
        try:
            cur.execute("UPDATE properties SET house_no=%s,street_id=%s,type_id=%s,status=%s WHERE id=%s",
                        (house_no, street_id, type_id, status, pid))
            self.conn.commit(); return True, "✅ Property updated successfully"
        except Error as e:
            self.conn.rollback(); return False, f"Error: {e}"
        finally: cur.close()

    def delete_property(self, property_id):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT COUNT(*) as c FROM payments WHERE property_id=%s", (property_id,))
            if cur.fetchone()['c'] > 0:
                return False, "⛔ Cannot delete property: payment records exist. Delete payments first."
            cur.execute("""SELECT r.id, r.name, COUNT(p.id) as pc FROM residents r
                           LEFT JOIN payments p ON r.id=p.resident_id
                           WHERE r.property_id=%s GROUP BY r.id HAVING COUNT(p.id)>0""", (property_id,))
            rwp = cur.fetchall()
            if rwp:
                names = [f"{r['name']} ({r['pc']} payments)" for r in rwp]
                return False, "⛔ Cannot delete: residents have payments:\n" + "\n".join(f"  • {n}" for n in names)
            cur.execute("SELECT COUNT(*) as c FROM residents WHERE property_id=%s AND is_active=1", (property_id,))
            if cur.fetchone()['c'] > 0:
                return False, "⛔ Cannot delete: active residents exist. Deactivate them first."
            cur.execute("SELECT COUNT(*) as c FROM residents WHERE property_id=%s", (property_id,))
            if cur.fetchone()['c'] > 0:
                return False, "⛔ Cannot delete: resident records exist. Delete them from Residents section first."
            cur.execute("DELETE FROM properties WHERE id=%s", (property_id,))
            self.conn.commit(); return True, "✅ Property deleted successfully"
        except Error as e:
            self.conn.rollback()
            msg = str(e)
            if "1451" in msg or "foreign key" in msg.lower():
                return False, "⛔ Database constraint error. Delete all payments and residents for this property first."
            return False, f"Error: {msg}"
        finally: cur.close()

    # ── Residents ────────────────────────────────────────────────
    def get_residents(self, include_inactive_only=False, include_all=False):
        cur = self.conn.cursor(dictionary=True)
        try:
            if include_inactive_only:
                where = "WHERE r.is_active=0"
            elif include_all:
                where = ""
            else:
                where = "WHERE r.is_active=1"
            cur.execute(f"""SELECT r.*, pr.house_no, s.name as street, pt.name as type_name,
                               b.total_due, b.total_paid, b.balance,
                               b.balance as current_balance
                            FROM residents r
                            LEFT JOIN properties pr ON r.property_id=pr.id
                            LEFT JOIN streets s ON pr.street_id=s.id
                            LEFT JOIN property_types pt ON pr.type_id=pt.id
                            LEFT JOIN balances b ON r.id=b.resident_id AND b.year=YEAR(CURDATE())
                            {where} ORDER BY s.name, r.name""")
            return cur.fetchall()
        finally: cur.close()

    def get_filtered_residents(self, payment_type='All', year='All',
                                occupancy='All', street='All',
                                status_filter='Active Only', house_type='All'):
        cur = self.conn.cursor(dictionary=True)
        try:
            year_int = int(year) if year != 'All' else None
            q = """SELECT r.*, pr.house_no, s.name as street, pt.name as type_name,
                          b.total_due, b.total_paid, b.balance, b.balance as current_balance,
                          CASE
                            WHEN b.balance IS NULL THEN 'No Record'
                            WHEN b.balance <= 0 THEN 'Fully Paid'
                            WHEN b.total_paid > 0 THEN 'Partial Payment'
                            ELSE 'No Payment'
                          END as payment_status,
                          lp.payment_date as last_payment_date,
                          lp.payment_description as last_payment_description
                   FROM residents r
                   LEFT JOIN properties pr ON r.property_id=pr.id
                   LEFT JOIN streets s ON pr.street_id=s.id
                   LEFT JOIN property_types pt ON pr.type_id=pt.id
                   LEFT JOIN balances b ON r.id=b.resident_id
                   LEFT JOIN payments lp ON lp.id = (
                       SELECT id FROM payments WHERE resident_id=r.id
                       ORDER BY payment_date DESC, id DESC LIMIT 1
                   )"""
            if year_int:
                q += " AND b.year=%s"
                params = [year_int]
            else:
                q += " AND (b.year=YEAR(CURDATE()) OR b.year IS NULL)"
                params = []
            conds = []
            if status_filter == 'Active Only': conds.append("r.is_active=1")
            elif status_filter == 'Inactive Only': conds.append("r.is_active=0")
            if occupancy != 'All': conds.append("r.occupancy_type=%s"); params.append(occupancy)
            if street != 'All': conds.append("s.name=%s"); params.append(street)
            if house_type != 'All': conds.append("pt.name=%s"); params.append(house_type)
            if payment_type != 'All':
                conds.append("""r.id IN (SELECT DISTINCT p2.resident_id FROM payments p2
                                         LEFT JOIN payment_types ptype ON p2.payment_type_id=ptype.id
                                         WHERE ptype.code=%s)""")
                params.append(payment_type)
            if conds: q += " WHERE " + " AND ".join(conds)
            q += " ORDER BY s.name, r.name"
            cur.execute(q, params)
            rows = cur.fetchall()
            return rows
        except Error as e:
            st.error(f"Filter error: {e}"); return []
        finally: cur.close()

    def get_resident_payment_history(self, resident_id):
        """Returns dict keyed by year with payment summary — matches original behaviour"""
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT DISTINCT service_year as year FROM payments WHERE resident_id=%s ORDER BY service_year DESC", (resident_id,))
            years = [r['year'] for r in cur.fetchall()]
            cur.execute("SELECT DISTINCT year FROM balances WHERE resident_id=%s ORDER BY year DESC", (resident_id,))
            bal_years = [r['year'] for r in cur.fetchall()]
            all_years = sorted(set(years + bal_years), reverse=True)
            history = {}
            for yr in all_years:
                cur.execute("SELECT SUM(amount_paid) as tp, COUNT(*) as pc FROM payments WHERE resident_id=%s AND service_year=%s", (resident_id, yr))
                pr = cur.fetchone()
                cur.execute("SELECT total_due, total_paid, balance, last_payment_date FROM balances WHERE resident_id=%s AND year=%s", (resident_id, yr))
                bi = cur.fetchone()
                outstanding_bf = 0
                if yr > min(all_years) if all_years else 0:
                    cur.execute("SELECT balance FROM balances WHERE resident_id=%s AND year=%s", (resident_id, yr-1))
                    pb = cur.fetchone()
                    if pb and pb['balance'] and pb['balance'] > 0:
                        outstanding_bf = float(pb['balance'])
                history[yr] = {
                    'total_paid': float(pr['tp'] or 0) if pr else 0,
                    'payment_count': int(pr['pc'] or 0) if pr else 0,
                    'balance': float(bi['balance'] or 0) if bi else 0,
                    'total_due': float(bi['total_due'] or 0) if bi else 240000,
                    'outstanding_brought_forward': outstanding_bf,
                    'last_payment_date': bi['last_payment_date'] if bi else None,
                }
            return history
        except Error as e:
            st.error(f"History error: {e}"); return {}
        finally: cur.close()

    def add_resident(self, name, property_id, phone, email, occupancy_type, join_date):
        cur = self.conn.cursor()
        try:
            cur.execute("""INSERT INTO residents(name,property_id,phone,email,occupancy_type,join_date,is_active)
                           VALUES(%s,%s,%s,%s,%s,%s,1)""",
                        (name, property_id, phone or None, email or None, occupancy_type, join_date))
            rid = cur.lastrowid
            due = 0.0 if occupancy_type == 'LANDLORD' else 240000.0
            cur.execute("""INSERT IGNORE INTO balances(resident_id,year,total_due,total_paid,balance)
                           VALUES(%s,YEAR(CURDATE()),%s,0,%s)""", (rid, due, due))
            if property_id:
                cur.execute("UPDATE properties SET status='Occupied' WHERE id=%s", (property_id,))
            self.conn.commit()
            return True, rid, "✅ Resident added successfully"
        except Error as e:
            self.conn.rollback(); return False, None, f"Error: {e}"
        finally: cur.close()

    def update_resident(self, rid, name, property_id, phone, email, occupancy_type):
        cur = self.conn.cursor()
        try:
            cur.execute("""UPDATE residents SET name=%s,property_id=%s,phone=%s,email=%s,occupancy_type=%s WHERE id=%s""",
                        (name, property_id, phone or None, email or None, occupancy_type, rid))
            self.conn.commit(); return True, "✅ Resident updated successfully"
        except Error as e:
            self.conn.rollback(); return False, f"Error: {e}"
        finally: cur.close()

    def delete_resident(self, resident_id, force_delete=False):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT COUNT(*) as c FROM payments WHERE resident_id=%s", (resident_id,))
            pc = cur.fetchone()['c']
            if pc > 0:
                return False, f"⛔ Cannot delete: {pc} payment record(s) exist. Delete payments first."
            cur.execute("SELECT COALESCE(SUM(balance),0) as t FROM balances WHERE resident_id=%s AND balance>0", (resident_id,))
            bal = float(cur.fetchone()['t'] or 0)
            if bal > 0 and not force_delete:
                return False, f"⛔ Cannot delete: outstanding balance of {format_currency(bal)}. Check 'Clear balance on deactivate' to proceed."
            if force_delete:
                cur.execute("DELETE FROM balances WHERE resident_id=%s", (resident_id,))
            cur.execute("SELECT property_id FROM residents WHERE id=%s", (resident_id,))
            row = cur.fetchone()
            prop_id = row['property_id'] if row else None
            cur.execute("UPDATE residents SET is_active=0 WHERE id=%s", (resident_id,))
            if prop_id:
                cur.execute("SELECT COUNT(*) as c FROM residents WHERE property_id=%s AND is_active=1", (prop_id,))
                if cur.fetchone()['c'] == 0:
                    cur.execute("UPDATE properties SET status='Vacant' WHERE id=%s", (prop_id,))
            self.conn.commit()
            if force_delete and bal > 0:
                return True, f"✅ Resident deactivated (cleared balance of {format_currency(bal)})"
            return True, "✅ Resident deactivated successfully"
        except Error as e:
            self.conn.rollback()
            msg = str(e)
            if "1451" in msg or "foreign key" in msg.lower():
                return False, "⛔ Cannot delete: resident has payment records. Delete payments first."
            return False, f"Error: {msg}"
        finally: cur.close()

    def reactivate_resident(self, resident_id):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT * FROM residents WHERE id=%s", (resident_id,))
            r = cur.fetchone()
            if not r: return False, "Resident not found"
            if r['is_active'] == 1: return False, "Resident is already active"
            cur.execute("UPDATE residents SET is_active=1 WHERE id=%s", (resident_id,))
            if r['property_id']:
                cur.execute("UPDATE properties SET status='Occupied' WHERE id=%s", (r['property_id'],))
            self.conn.commit(); return True, "✅ Resident reactivated successfully"
        except Error as e:
            self.conn.rollback(); return False, f"Error: {e}"
        finally: cur.close()

    def search_residents(self, term):
        cur = self.conn.cursor(dictionary=True)
        try:
            like = f"%{term}%"
            cur.execute("""SELECT r.*, pr.house_no, s.name as street,
                                  b.total_due, b.total_paid, b.balance, b.balance as current_balance
                           FROM residents r
                           LEFT JOIN properties pr ON r.property_id=pr.id
                           LEFT JOIN streets s ON pr.street_id=s.id
                           LEFT JOIN balances b ON r.id=b.resident_id AND b.year=YEAR(CURDATE())
                           WHERE r.is_active=1 AND (r.name LIKE %s OR r.phone LIKE %s OR r.email LIKE %s)
                           ORDER BY r.name LIMIT 20""", (like, like, like))
            return cur.fetchall()
        finally: cur.close()

    # ── Payments ─────────────────────────────────────────────────
    def _update_balance(self, cur, resident_id, year, amount_paid, total_due, payment_date):
        cur.execute("SELECT * FROM balances WHERE resident_id=%s AND year=%s", (resident_id, year))
        b = cur.fetchone()
        if b:
            # Use existing total_due if valid (>0), otherwise fall back to the passed-in value
            effective_due = float(b['total_due']) if float(b['total_due'] or 0) > 0 else float(total_due)
            new_paid = float(b['total_paid']) + float(amount_paid)
            new_bal = effective_due - new_paid
            cur.execute("""UPDATE balances SET total_due=%s,total_paid=%s,balance=%s,last_payment_date=%s
                           WHERE resident_id=%s AND year=%s""",
                        (effective_due, new_paid, new_bal, payment_date, resident_id, year))
        else:
            new_bal = float(total_due) - float(amount_paid)
            cur.execute("""INSERT INTO balances(resident_id,year,total_due,total_paid,balance,last_payment_date)
                           VALUES(%s,%s,%s,%s,%s,%s)""",
                        (resident_id, year, total_due, amount_paid, new_bal, payment_date))

    def update_balance_after_payment(self, resident_id, year, amount_paid, payment_date):
        cur = self.conn.cursor(dictionary=True)
        try: self._update_balance(cur, resident_id, year, amount_paid, 240000.0, payment_date)
        finally: cur.close()

    def add_payment(self, resident_id, property_id, amount_paid, payment_date,
                    service_year, payment_type_id, payment_method_id,
                    bank_account_id, payment_description):
        cur = self.conn.cursor(dictionary=True)
        try:
            if float(amount_paid) <= 0:
                return False, None, "Payment amount must be greater than zero"
            amount_paid = Decimal(str(amount_paid))
            cur.execute("SELECT default_amount FROM payment_types WHERE id=%s", (payment_type_id,))
            pt = cur.fetchone()
            default_due = Decimal(str(pt['default_amount'])) if pt else Decimal('240000')
            cur.execute("SELECT total_due,total_paid FROM balances WHERE resident_id=%s AND year=%s",
                        (resident_id, service_year))
            bal = cur.fetchone()
            if bal:
                # Use existing total_due if valid (>0), otherwise fall back to default
                existing_due = Decimal(str(bal['total_due'] or 0))
                total_due = existing_due if existing_due > 0 else default_due
                current_paid = Decimal(str(bal['total_paid'] or 0))
            else:
                total_due = default_due; current_paid = Decimal('0')
            balance_due = total_due - (current_paid + amount_paid)
            today = datetime.now()
            pfx = today.strftime("%Y-%m-%d")
            cur.execute("SELECT COUNT(*) as c FROM payments WHERE receipt_no LIKE %s", (f"REC-{pfx}-%",))
            n = cur.fetchone()['c'] + 1
            receipt_no = f"REC-{pfx}-{n:03d}"
            cur.execute("""INSERT INTO payments(receipt_no,receipt_issued_date,payment_date,
                               resident_id,property_id,amount_due,amount_paid,balance_due,
                               bank_account_id,payment_method_id,payment_type_id,
                               payment_description,service_year)
                           VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (receipt_no, today.date(), payment_date, resident_id, property_id,
                         float(total_due), float(amount_paid), float(balance_due),
                         bank_account_id, payment_method_id, payment_type_id,
                         payment_description, service_year))
            self._update_balance(cur, resident_id, service_year,
                                  float(amount_paid), float(total_due), payment_date)
            self.conn.commit()
            return True, receipt_no, "Payment recorded successfully"
        except Error as e:
            self.conn.rollback(); return False, None, f"Error: {e}"
        finally: cur.close()

    def update_payment(self, pid, amount_paid, payment_date, payment_type_id,
                       payment_method_id, bank_account_id, description, service_year):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT * FROM payments WHERE id=%s", (pid,))
            old = cur.fetchone()
            if not old: return False, "Payment not found"
            diff = float(amount_paid) - float(old['amount_paid'])
            cur.execute("""UPDATE payments SET amount_paid=%s,payment_date=%s,payment_type_id=%s,
                           payment_method_id=%s,bank_account_id=%s,payment_description=%s,
                           service_year=%s WHERE id=%s""",
                        (amount_paid, payment_date, payment_type_id, payment_method_id,
                         bank_account_id, description, service_year, pid))
            if diff != 0:
                cur.execute("""UPDATE balances SET total_paid=total_paid+%s,balance=balance-%s
                               WHERE resident_id=%s AND year=%s""",
                            (diff, diff, old['resident_id'], service_year))
            self.conn.commit(); return True, "✅ Payment updated successfully"
        except Error as e:
            self.conn.rollback(); return False, f"Error: {e}"
        finally: cur.close()

    def delete_payment(self, payment_id):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT * FROM payments WHERE id=%s", (payment_id,))
            p = cur.fetchone()
            if not p: return False, "Payment not found"
            cur.execute("""UPDATE balances SET total_paid=total_paid-%s,balance=balance+%s
                           WHERE resident_id=%s AND year=%s""",
                        (p['amount_paid'], p['amount_paid'], p['resident_id'], p['service_year']))
            cur.execute("DELETE FROM payments WHERE id=%s", (payment_id,))
            self.conn.commit(); return True, "✅ Payment deleted successfully"
        except Error as e:
            self.conn.rollback(); return False, f"Error: {e}"
        finally: cur.close()

    def carry_forward_outstanding(self, resident_id, from_year, to_year):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT balance,total_due FROM balances WHERE resident_id=%s AND year=%s",
                        (resident_id, from_year))
            prev = cur.fetchone()
            if not prev or float(prev['balance'] or 0) <= 0:
                return False, "No outstanding balance to carry forward"
            outstanding = Decimal(str(prev['balance']))
            new_total_due = Decimal('240000') + outstanding
            cur.execute("SELECT * FROM balances WHERE resident_id=%s AND year=%s", (resident_id, to_year))
            cy = cur.fetchone()
            if cy:
                current_paid = Decimal(str(cy['total_paid']))
                new_bal = new_total_due - current_paid
                cur.execute("UPDATE balances SET total_due=%s,balance=%s WHERE resident_id=%s AND year=%s",
                            (float(new_total_due), float(new_bal), resident_id, to_year))
            else:
                cur.execute("INSERT INTO balances(resident_id,year,total_due,total_paid,balance) VALUES(%s,%s,%s,0,%s)",
                            (resident_id, to_year, float(new_total_due), float(new_total_due)))
            self.conn.commit()
            return True, f"Outstanding {format_currency(float(outstanding))} carried forward to {to_year}"
        except Error as e:
            self.conn.rollback(); return False, f"Error: {e}"
        finally: cur.close()

    def get_resident_payments(self, resident_id):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("""SELECT p.*, pt.code, pt.name as payment_type
                           FROM payments p LEFT JOIN payment_types pt ON p.payment_type_id=pt.id
                           WHERE p.resident_id=%s ORDER BY p.payment_date DESC""", (resident_id,))
            return cur.fetchall()
        finally: cur.close()

    def get_receipt(self, receipt_no):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("""SELECT p.*, r.name, r.phone, r.email,
                                  pr.house_no, s.name as street,
                                  pt.name as payment_type, pt.code as payment_code,
                                  pm.name as payment_method,
                                  ba.bank_name, ba.account_number
                           FROM payments p
                           JOIN residents r ON p.resident_id=r.id
                           LEFT JOIN properties pr ON p.property_id=pr.id
                           LEFT JOIN streets s ON pr.street_id=s.id
                           LEFT JOIN payment_types pt ON p.payment_type_id=pt.id
                           LEFT JOIN payment_methods pm ON p.payment_method_id=pm.id
                           LEFT JOIN bank_accounts ba ON p.bank_account_id=ba.id
                           WHERE p.receipt_no=%s""", (receipt_no,))
            return cur.fetchone()
        finally: cur.close()

    # ── Reports ──────────────────────────────────────────────────
    def get_payments_by_date_range(self, from_date, to_date,
                                    occupancy='All', payment_type='All',
                                    street='All', year='All',
                                    resident_name='', house_no='',
                                    house_type='All', phone='', email='',
                                    amount_min=None, amount_max=None):
        cur = self.conn.cursor(dictionary=True)
        try:
            q = """SELECT p.*, r.name as resident_name, r.phone, r.email, r.occupancy_type,
                          pr.house_no, s.name as street, ptype2.name as house_type,
                          pt.code as payment_code, pt.name as payment_type,
                          pm.name as payment_method,
                          ba.bank_name, ba.account_number
                   FROM payments p
                   JOIN residents r ON p.resident_id=r.id
                   LEFT JOIN properties pr ON p.property_id=pr.id
                   LEFT JOIN streets s ON pr.street_id=s.id
                   LEFT JOIN property_types ptype2 ON pr.type_id=ptype2.id
                   LEFT JOIN payment_types pt ON p.payment_type_id=pt.id
                   LEFT JOIN payment_methods pm ON p.payment_method_id=pm.id
                   LEFT JOIN bank_accounts ba ON p.bank_account_id=ba.id
                   WHERE p.payment_date BETWEEN %s AND %s"""
            params = [from_date, to_date]
            if occupancy != 'All': q += " AND r.occupancy_type=%s"; params.append(occupancy)
            if payment_type != 'All': q += " AND pt.code=%s"; params.append(payment_type)
            if street != 'All': q += " AND s.name=%s"; params.append(street)
            if year != 'All': q += " AND p.service_year=%s"; params.append(int(year))
            if resident_name: q += " AND r.name LIKE %s"; params.append(f"%{resident_name}%")
            if house_no: q += " AND pr.house_no LIKE %s"; params.append(f"%{house_no}%")
            if house_type != 'All': q += " AND ptype2.name=%s"; params.append(house_type)
            if phone: q += " AND r.phone LIKE %s"; params.append(f"%{phone}%")
            if email: q += " AND r.email LIKE %s"; params.append(f"%{email}%")
            if amount_min is not None: q += " AND p.amount_paid >= %s"; params.append(float(amount_min))
            if amount_max is not None: q += " AND p.amount_paid <= %s"; params.append(float(amount_max))
            q += " ORDER BY p.payment_date DESC, p.id DESC"
            cur.execute(q, params); return cur.fetchall()
        finally: cur.close()

    def get_payment_summary_by_type(self, year=None):
        cur = self.conn.cursor(dictionary=True)
        try:
            where = f"WHERE p.service_year={year}" if year else ""
            cur.execute(f"""SELECT pt.code, pt.name, COUNT(*) as cnt,
                               SUM(p.amount_paid) as total_paid
                           FROM payments p LEFT JOIN payment_types pt ON p.payment_type_id=pt.id
                           {where} GROUP BY pt.id,pt.code,pt.name ORDER BY total_paid DESC""")
            return cur.fetchall()
        finally: cur.close()

    # ── Lookup helpers ───────────────────────────────────────────
    def get_streets(self):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT DISTINCT name FROM streets ORDER BY name")
            return cur.fetchall()
        finally: cur.close()

    def get_all_streets(self):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT * FROM streets ORDER BY name")
            return cur.fetchall()
        finally: cur.close()

    def get_property_types(self):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT id, name FROM property_types ORDER BY name")
            return cur.fetchall()
        finally: cur.close()

    def get_payment_types(self):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT * FROM payment_types ORDER BY code")
            return cur.fetchall()
        finally: cur.close()

    def get_payment_methods(self):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT * FROM payment_methods ORDER BY name")
            return cur.fetchall()
        finally: cur.close()

    def get_bank_accounts(self):
        cur = self.conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT * FROM bank_accounts WHERE is_active=1 ORDER BY bank_name")
            return cur.fetchall()
        finally: cur.close()

    # ── Settings CRUD ────────────────────────────────────────────
    def upsert_payment_type(self, code, name, description, freq, amount):
        cur = self.conn.cursor()
        try:
            cur.execute("""INSERT INTO payment_types(code,name,description,charge_frequency,default_amount)
                           VALUES(%s,%s,%s,%s,%s)
                           ON DUPLICATE KEY UPDATE name=%s,description=%s,charge_frequency=%s,default_amount=%s""",
                        (code,name,description,freq,amount,name,description,freq,amount))
            self.conn.commit(); return True, "✅ Saved"
        except Error as e:
            self.conn.rollback(); return False, str(e)
        finally: cur.close()

    def upsert_payment_method(self, name, description):
        cur = self.conn.cursor()
        try:
            cur.execute("INSERT INTO payment_methods(name,description) VALUES(%s,%s) ON DUPLICATE KEY UPDATE description=%s",
                        (name,description,description))
            self.conn.commit(); return True, "✅ Saved"
        except Error as e:
            self.conn.rollback(); return False, str(e)
        finally: cur.close()

    def upsert_bank_account(self, bank_name, account_name, account_number, branch):
        cur = self.conn.cursor()
        try:
            cur.execute("""INSERT INTO bank_accounts(bank_name,account_name,account_number,branch,is_active)
                           VALUES(%s,%s,%s,%s,1)
                           ON DUPLICATE KEY UPDATE bank_name=%s,account_name=%s,branch=%s""",
                        (bank_name,account_name,account_number,branch,bank_name,account_name,branch))
            self.conn.commit(); return True, "✅ Saved"
        except Error as e:
            self.conn.rollback(); return False, str(e)
        finally: cur.close()

    def delete_bank_account(self, aid):
        cur = self.conn.cursor()
        try:
            cur.execute("UPDATE bank_accounts SET is_active=0 WHERE id=%s", (aid,))
            self.conn.commit(); return True, "Account deactivated"
        except Error as e:
            self.conn.rollback(); return False, str(e)
        finally: cur.close()

    def clear_all_tables(self):
        cur = self.conn.cursor()
        try:
            cur.execute("SET FOREIGN_KEY_CHECKS=0")
            for t in ['balances','payments','residents','properties','streets',
                      'property_types','payment_types','payment_methods','bank_accounts']:
                try:
                    cur.execute(f"TRUNCATE TABLE {t}")
                except: pass
            cur.execute("SET FOREIGN_KEY_CHECKS=1")
            self.conn.commit(); return True, "All tables cleared successfully"
        except Error as e:
            self.conn.rollback(); return False, str(e)
        finally: cur.close()

    # ── Excel template upload ──────────────────────────────────────
    # Reads the estate Excel file and distributes data into tables:
    #   STREET column           → streets table
    #   HOUSE NO. + HOUSE TYPE  → properties table
    #   RESIDENT'S NAME + ...   → residents table
    #   PAYMENT > 0 rows        → payments table
    #   BFWD rows               → balances table (opening balances)
    # Does NOT store the file itself — only extracts and saves the text data.
    def process_excel_template_upload(self, df):
        import traceback as _tb
        import pandas as _pd
        results = {
            'success': 0, 'errors': [], 'warnings': [], 'skipped': 0,
            'streets_created': 0, 'properties_created': 0,
            'residents_created': 0, 'payments_created': 0, 'balances_created': 0
        }
        cur = self.conn.cursor(dictionary=True)

        # Normalise all column names to uppercase
        df = df.copy()
        df.columns = [str(c).strip().upper() for c in df.columns]

        def fcol(aliases):
            for a in aliases:
                if a in df.columns: return a
            return None

        C = {
            'DATE':    fcol(['DATE', 'RECEIPT DATE']),
            'NAME':    fcol(["RESIDENT'S NAME", 'RESIDENTS NAME', 'NAME', 'RESIDENT NAME']),
            'RECEIPT': fcol(['RECEIPT NO.', 'RECEIPT NO', 'RECEIPT_NO']),
            'PAID_ON': fcol(['PAID ON', 'PAYMENT DATE', 'PAID_ON']),
            'HOUSE':   fcol(['HOUSE NO.', 'HOUSE NO', 'HOUSE_NO']),
            'STREET':  fcol(['STREET']),
            'TYPE':    fcol(['HOUSE TYPE', 'HOUSE_TYPE', 'TYPE']),
            'OCC':     fcol(['OCCUPANCY', 'OCCUPANCY TYPE']),
            'PHONE':   fcol(['PHONE NO.', 'PHONE NO', 'PHONE']),
            'EMAIL':   fcol(['EMAIL']),
            'DESC':    fcol(['DESCRIPTION', 'DESC']),
            'PAYMENT': fcol(['PAYMENT']),
            'SVC':     fcol(['SVC STATUS', 'SERVICE CHARGE', 'SVC']),
            'INFRA':   fcol(['INFRA.', 'INFRA']),
            'LEGAL':   fcol(['LEGAL']),
            'TRANSF':  fcol(['TRANSF.', 'TRANSF', 'TRANSFORMER']),
            'DEV':     fcol(['DEV. LEVY', 'DEV.LEVY', 'DEV LEVY', 'DEV']),
            'PARTY':   fcol(['YR. END PARTY', 'YR END PARTY', 'PARTY']),
        }

        def cv(key, row):
            col = C.get(key)
            if col and col in row.index:
                val = row[col]
                if val is None: return ''
                s = str(val).strip()
                return '' if s.upper() in ('NAN', 'NONE', 'NAT') else s
            return ''

        def cfloat(key, row):
            col = C.get(key)
            if not col or col not in row.index: return 0.0
            try:
                v = row[col]
                if v is None or str(v).strip().upper() in ('NAN', '', 'NONE'): return 0.0
                return float(v)
            except: return 0.0

        def parse_date(raw):
            if raw is None: return None
            s = str(raw).strip().upper()
            if s in ('NAN', '', 'NONE', 'BFWD', 'PMT', 'B/FWD', 'NAT'): return None
            try: return _pd.to_datetime(raw).date()
            except: return None

        VALID_OCC = {'LANDLORD', 'RESIDENT', 'TENANT', 'UNOCCUPIED', 'UNCOMPLETED', 'UNDEVELOPED'}

        try:
            # ── Seed lookup tables ────────────────────────────────
            for code, info in SERVICE_CHARGES.items():
                try:
                    cur.execute(
                        "INSERT IGNORE INTO payment_types(code,name,charge_frequency,default_amount) VALUES(%s,%s,%s,%s)",
                        (code, info['name'], info['freq'], info['amount'])
                    )
                except: pass
            for mn, md in [('Cash','Cash payment'),('Bank Transfer','Electronic bank transfer'),
                           ('Cheque','Cheque payment'),('POS','Point of sale')]:
                try: cur.execute("INSERT IGNORE INTO payment_methods(name,description) VALUES(%s,%s)", (mn, md))
                except: pass
            for tn in ['Bungalow','Duplex','Apartment','Shopping Complex']:
                try: cur.execute("INSERT IGNORE INTO property_types(name) VALUES(%s)", (tn,))
                except: pass
            self.conn.commit()

            # ── Pre-load ALL lookup data into Python dicts (avoids per-row SELECT) ──
            cur.execute("SELECT id, code FROM payment_types")
            pt_map = {r['code']: r['id'] for r in cur.fetchall()}

            cur.execute("SELECT id FROM payment_methods WHERE name='Bank Transfer' LIMIT 1")
            pm_row = cur.fetchone()
            default_pm_id = pm_row['id'] if pm_row else None

            cur.execute("SELECT id, UPPER(name) as n FROM streets")
            street_cache = {r['n']: r['id'] for r in cur.fetchall()}

            cur.execute("SELECT id, UPPER(name) as n FROM property_types")
            type_cache = {r['n']: r['id'] for r in cur.fetchall()}

            cur.execute("SELECT id, house_no, street_id FROM properties")
            prop_cache = {(r['house_no'], r['street_id']): r['id'] for r in cur.fetchall()}

            cur.execute("SELECT id, name, property_id FROM residents")
            res_cache = {(r['name'], r['property_id']): r['id'] for r in cur.fetchall()}

            cur.execute("SELECT receipt_no FROM payments")
            receipt_cache = {r['receipt_no'] for r in cur.fetchall()}

            # Balance cache: (resident_id, year) -> {total_due, total_paid}
            cur.execute("SELECT resident_id, year, total_due, total_paid FROM balances")
            bal_cache = {(r['resident_id'], r['year']): {'total_due': float(r['total_due']), 'total_paid': float(r['total_paid'])} for r in cur.fetchall()}

            # Auto-receipt counter cache: date_prefix -> count
            cur.execute("SELECT receipt_no FROM payments WHERE receipt_no LIKE 'REC-%'")
            auto_receipt_counter = {}
            for r in cur.fetchall():
                parts = r['receipt_no'].split('-')
                if len(parts) >= 4:
                    pfx = '-'.join(parts[1:4])
                    auto_receipt_counter[pfx] = auto_receipt_counter.get(pfx, 0) + 1

            today = date.today()

            # ── Process each row ─────────────────────────────────
            for idx, row in df.iterrows():
                try:
                    name = cv('NAME', row)
                    if not name or name.upper() in ('NAN', '', "RESIDENT'S NAME", 'NAME', 'RESIDENTS NAME'):
                        continue
                    if name.upper() in ('CANCELLED', 'VOID', 'CANCEL', 'N/A', 'DELETED'):
                        results['warnings'].append(f"Row {idx+2}: Skipped — marked as '{name}'")
                        continue

                    receipt_raw  = cv('RECEIPT', row)
                    receipt_flag = receipt_raw.upper().strip()
                    is_bfwd = receipt_flag in ('BFWD', 'B/FWD')

                    house_no    = cv('HOUSE', row).strip().upper()
                    street_name = cv('STREET', row).strip().upper()

                    if not street_name or street_name in ('NAN', '-', ''):
                        results['skipped'] += 1; continue
                    if not house_no or house_no in ('NAN', ''):
                        results['skipped'] += 1; continue

                    house_type = cv('TYPE', row).strip().title() or 'Bungalow'
                    occ_raw    = cv('OCC', row).strip().upper()
                    if occ_raw and occ_raw not in VALID_OCC and occ_raw not in ('NAN','','NONE'):
                        results['warnings'].append(
                            f"Row {idx+2} ({name}): Occupancy '{occ_raw}' not recognised — saved as TENANT")
                    occupancy  = occ_raw if occ_raw in VALID_OCC else 'TENANT'
                    phone      = cv('PHONE', row) or None
                    email      = cv('EMAIL', row) or None
                    description = cv('DESC', row)

                    date_col   = C.get('DATE')
                    paidon_col = C.get('PAID_ON')
                    receipt_date = parse_date(row[date_col]   if date_col   and date_col   in row.index else None)
                    paid_on      = parse_date(row[paidon_col] if paidon_col and paidon_col in row.index else None)
                    if not receipt_date: receipt_date = paid_on or today
                    if not paid_on:      paid_on      = receipt_date or today
                    service_year = receipt_date.year

                    payment_amt = cfloat('PAYMENT', row)
                    svc_val     = cfloat('SVC', row)
                    infra_val   = cfloat('INFRA',  row)
                    legal_val   = cfloat('LEGAL',  row)
                    transf_val  = cfloat('TRANSF', row)
                    dev_val     = cfloat('DEV',    row)
                    party_val   = cfloat('PARTY',  row)

                    # ── streets (cache-first, insert only if new) ──
                    if street_name not in street_cache:
                        cur.execute("INSERT INTO streets(name) VALUES(%s)", (street_name,))
                        street_cache[street_name] = cur.lastrowid
                        results['streets_created'] += 1
                    street_id = street_cache[street_name]

                    # ── property_types (cache-first) ──
                    ht_key = house_type.upper()
                    if ht_key not in type_cache:
                        try:
                            cur.execute("INSERT INTO property_types(name) VALUES(%s)", (house_type,))
                            type_cache[ht_key] = cur.lastrowid
                        except: type_cache[ht_key] = list(type_cache.values())[0] if type_cache else 1
                    type_id = type_cache[ht_key]

                    # ── properties (cache-first) ──
                    prop_key = (house_no, street_id)
                    if prop_key not in prop_cache:
                        cur.execute(
                            "INSERT INTO properties(house_no,street_id,type_id,status) VALUES(%s,%s,%s,'Occupied')",
                            (house_no, street_id, type_id)
                        )
                        prop_cache[prop_key] = cur.lastrowid
                        results['properties_created'] += 1
                    prop_id = prop_cache[prop_key]

                    # ── residents (cache-first) ──
                    res_key = (name, prop_id)
                    if res_key not in res_cache:
                        cur.execute(
                            "INSERT INTO residents(name,property_id,phone,email,occupancy_type,is_active,join_date) VALUES(%s,%s,%s,%s,%s,1,%s)",
                            (name, prop_id, phone, email, occupancy, receipt_date)
                        )
                        res_id = cur.lastrowid
                        res_cache[res_key] = res_id
                        results['residents_created'] += 1
                    else:
                        res_id = res_cache[res_key]
                        if phone or email:
                            cur.execute(
                                "UPDATE residents SET phone=COALESCE(%s,phone),email=COALESCE(%s,email),occupancy_type=%s WHERE id=%s",
                                (phone, email, occupancy, res_id)
                            )

                    # ── balances for BFWD rows ──
                    if is_bfwd:
                        bal_amount = 0.0 if occupancy == 'LANDLORD' else (abs(svc_val) if svc_val != 0 else 240000.0)
                        cur.execute(
                            """INSERT INTO balances(resident_id,year,total_due,total_paid,balance)
                               VALUES(%s,%s,%s,0,%s)
                               ON DUPLICATE KEY UPDATE total_due=%s, balance=%s""",
                            (res_id, service_year, bal_amount, bal_amount, bal_amount, bal_amount)
                        )
                        bal_cache[(res_id, service_year)] = {'total_due': bal_amount, 'total_paid': 0.0}
                        results['balances_created'] += 1
                        results['success'] += 1
                        continue

                    # ── payments ──
                    service_entries = []
                    if payment_amt > 0:  service_entries.append(('SVC',   payment_amt))
                    if infra_val  > 0:   service_entries.append(('INFRA',  infra_val))
                    if legal_val  > 0:   service_entries.append(('LEGAL',  legal_val))
                    if transf_val > 0:   service_entries.append(('TRANSF', transf_val))
                    if dev_val    > 0:   service_entries.append(('DEV',    dev_val))
                    if party_val  > 0:   service_entries.append(('PARTY',  party_val))

                    if not service_entries:
                        if receipt_flag not in ('','NAN','NONE','PMT','BFWD','B/FWD'):
                            results['warnings'].append(
                                f"Row {idx+2} ({name}): Receipt '{receipt_raw}' has no payment amounts — skipped")
                        results['skipped'] += 1
                        continue

                    for svc_code, amt in service_entries:
                        pt_id = pt_map.get(svc_code, pt_map.get('SVC', 1))
                        default_due = SERVICE_CHARGES.get(svc_code, {}).get('amount', 240000.0)
                        if occupancy == 'LANDLORD' and SERVICE_CHARGES.get(svc_code, {}).get('freq') == 'Annual':
                            default_due = 0.0

                        # Use cached balance (no DB query)
                        b_cached     = bal_cache.get((res_id, service_year))
                        total_due_v  = float(b_cached['total_due'])  if b_cached and float(b_cached['total_due'])  > 0 else default_due
                        current_paid = float(b_cached['total_paid']) if b_cached else 0.0
                        balance_after = total_due_v - (current_paid + amt)

                        # Build receipt number (no DB query for auto-receipts)
                        if receipt_raw and receipt_flag not in ('BFWD', 'PMT', 'NAN', ''):
                            r_no = f"REC-{receipt_raw}-{svc_code}"
                        else:
                            pfx = paid_on.strftime("%Y-%m-%d")
                            n = auto_receipt_counter.get(pfx, 0) + 1
                            auto_receipt_counter[pfx] = n
                            r_no = f"REC-{pfx}-{n:03d}-{svc_code}"

                        # Skip duplicates (cache-first)
                        if r_no in receipt_cache:
                            results['warnings'].append(
                                f"Row {idx+2} ({name}): Receipt {r_no} already saved — skipped")
                            continue
                        receipt_cache.add(r_no)

                        cur.execute(
                            """INSERT INTO payments(
                                   receipt_no, receipt_issued_date, payment_date,
                                   resident_id, property_id,
                                   amount_due, amount_paid, balance_due,
                                   payment_method_id, payment_type_id,
                                   payment_description, service_year)
                               VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (r_no, receipt_date, paid_on,
                             res_id, prop_id,
                             total_due_v, amt, balance_after,
                             default_pm_id, pt_id,
                             description or f"{svc_code} payment",
                             service_year)
                        )

                        # Update balance cache and DB
                        new_paid = current_paid + amt
                        bal_cache[(res_id, service_year)] = {'total_due': total_due_v, 'total_paid': new_paid}
                        self._update_balance(cur, res_id, service_year, amt, total_due_v, paid_on)
                        results['payments_created'] += 1
                        results['success'] += 1

                except Exception as row_err:
                    err_msg = str(row_err)
                    if any(x in err_msg.lower() for x in ("unknown column", "doesn't exist",
                           "table", "no such", "column")):
                        results['warnings'].append(
                            f"Row {idx+2} ({name}): DB column/table not available — {err_msg[:120]}")
                    else:
                        results['errors'].append(f"Row {idx+2} ({name}): {err_msg[:200]}")

            self.conn.commit()

        except Exception as e:
            self.conn.rollback()
            results['errors'].append(f"Upload failed: {str(e)}")
        finally:
            cur.close()

        return results



# ── PDF Generator ────────────────────────────────────────────────
def generate_pdf(receipt_data):
    # FPDF uses latin-1 internally — sanitize all text going in
    def s(text):
        return str(text).replace('\u2014', '-').replace('\u2013', '-')\
                        .replace('\u2019', "'").replace('\u2018', "'")\
                        .replace('\u201c', '"').replace('\u201d', '"')\
                        .encode('latin-1', errors='replace').decode('latin-1')

    if not receipt_data:
        pdf = FPDF(); pdf.add_page(); pdf.set_font("Arial","B",14)
        pdf.cell(0,10,"Receipt Not Found",ln=True,align="C")
        out = pdf.output(dest='S')
        return bytes(out) if isinstance(out,bytearray) else out.encode('latin-1')

    pdf = FPDF(); pdf.add_page()
    pdf.set_fill_color(16,185,129); pdf.rect(0,0,210,40,'F')
    pdf.set_font("Arial","B",18); pdf.set_text_color(255,255,255)
    pdf.cell(0,15,"SUNSHINE ESTATE",ln=True,align="C")
    pdf.set_font("Arial","",11)
    pdf.cell(0,8,"Estate Management System",ln=True,align="C")
    pdf.cell(0,8,"Payment Receipt",ln=True,align="C")
    pdf.set_text_color(0,0,0); pdf.ln(5)
    pdf.set_fill_color(16,185,129); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",11)
    pdf.cell(0,8,s(f"   RECEIPT NO: {receipt_data.get('receipt_no','N/A')}"),ln=True,fill=True)
    pdf.set_text_color(0,0,0); pdf.set_font("Arial","",10); pdf.ln(5)
    cw=95
    rd = receipt_data.get('receipt_issued_date'); pd_ = receipt_data.get('payment_date')
    pdf.cell(cw,6,s(f"Receipt Date: {rd.strftime('%d/%m/%Y') if rd and hasattr(rd,'strftime') else 'N/A'}"))
    pdf.cell(cw,6,s(f"Payment Date: {pd_.strftime('%d/%m/%Y') if pd_ and hasattr(pd_,'strftime') else 'N/A'}"),ln=True)
    pdf.cell(0,6,s(f"Service Year: {receipt_data.get('service_year','N/A')}"),ln=True); pdf.ln(5)
    pdf.set_fill_color(248,250,252); pdf.set_font("Arial","B",10)
    pdf.cell(0,7,"  RESIDENT INFORMATION",ln=True,fill=True); pdf.set_font("Arial","",10)
    pdf.cell(45,6,"Name:"); pdf.set_font("Arial","B",10)
    pdf.cell(0,6,s(safe_str(receipt_data.get('name'),'N/A')),ln=True); pdf.set_font("Arial","",10)
    pdf.cell(45,6,"Phone:"); pdf.cell(0,6,s(safe_str(receipt_data.get('phone'),'N/A')),ln=True)
    if receipt_data.get('email'):
        pdf.cell(45,6,"Email:"); pdf.cell(0,6,s(safe_str(receipt_data.get('email'))),ln=True)
    if receipt_data.get('house_no'):
        pdf.cell(45,6,"Property:")
        pdf.cell(0,6,s(f"{receipt_data.get('house_no','N/A')}, {receipt_data.get('street','N/A')}"),ln=True)
    pdf.ln(5)
    pdf.set_fill_color(248,250,252); pdf.set_font("Arial","B",10)
    pdf.cell(0,7,"  PAYMENT DETAILS",ln=True,fill=True); pdf.set_font("Arial","",10)
    code=receipt_data.get('payment_code','N/A'); ptype=receipt_data.get('payment_type','N/A')
    pdf.cell(45,6,"Payment Type:"); pdf.cell(0,6,s(f"{code} - {ptype}"),ln=True)
    if receipt_data.get('payment_method'):
        pdf.cell(45,6,"Payment Method:"); pdf.cell(0,6,s(safe_str(receipt_data.get('payment_method'))),ln=True)
    if receipt_data.get('bank_name'):
        pdf.cell(45,6,"Bank Account:")
        pdf.cell(0,6,s(f"{receipt_data['bank_name']} - {receipt_data.get('account_number','')}"),ln=True)
    if receipt_data.get('payment_description'):
        pdf.cell(45,6,"Description:"); pdf.cell(0,6,s(str(receipt_data['payment_description'])[:60]),ln=True)
    pdf.ln(3)
    pdf.set_draw_color(16,185,129); pdf.set_line_width(0.5)
    pdf.set_fill_color(16,185,129); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",10)
    pdf.cell(100,8,"Description",border=1,fill=True)
    pdf.cell(90,8,"Amount",ln=True,border=1,fill=True,align="R")
    pdf.set_text_color(0,0,0); pdf.set_font("Arial","",10); pdf.set_fill_color(255,255,255)
    pdf.cell(100,7,"Amount Due",border=1)
    pdf.cell(90,7,s(format_currency_pdf(receipt_data.get('amount_due',0))),ln=True,border=1,align="R")
    pdf.cell(100,7,"Amount Paid",border=1); pdf.set_font("Arial","B",10)
    pdf.cell(90,7,s(format_currency_pdf(receipt_data.get('amount_paid',0))),ln=True,border=1,align="R")
    pdf.set_font("Arial","",10)
    bal=receipt_data.get('balance_due',0)
    pdf.set_fill_color(254,226,226) if float(bal)>0 else pdf.set_fill_color(220,252,231)
    pdf.set_font("Arial","B",10)
    pdf.cell(100,8,"Balance Due",border=1,fill=True)
    pdf.cell(90,8,s(format_currency_pdf(bal)),ln=True,border=1,fill=True,align="R")
    pdf.ln(10); pdf.set_font("Arial","I",9); pdf.set_text_color(100,100,100)
    pdf.multi_cell(0,5,"Thank you for your payment. Please keep this receipt for your records.\nFor inquiries, contact the Estate Management Office.")
    pdf.ln(5); pdf.set_draw_color(0,0,0)
    pdf.line(20,pdf.get_y()+10,90,pdf.get_y()+10); pdf.ln(12)
    pdf.set_font("Arial","",9); pdf.cell(70,5,"Authorized Signature",align="C")
    out = pdf.output(dest='S')
    return bytes(out) if isinstance(out,bytearray) else out.encode('latin-1')


def generate_report_excel(payments, title="Payment Report"):
    wb = Workbook(); ws = wb.active; ws.title="Report"
    green="10B981"; white="FFFFFF"; light="F0FDF4"
    hf = Font(bold=True,color=white,name="Arial")
    hfill = PatternFill("solid",fgColor=green)
    thin = Border(left=Side(style='thin'),right=Side(style='thin'),
                  top=Side(style='thin'),bottom=Side(style='thin'))
    ws.merge_cells('A1:N1'); ws['A1']=title
    ws['A1'].font=Font(bold=True,size=14,color=green,name="Arial")
    ws['A1'].alignment=Alignment(horizontal='center'); ws.row_dimensions[1].height=25
    ws.append(['']*14)
    headers=['Receipt No','Date','Resident','Phone','Occupancy',
             'House No','Street','Payment Type','Method','Bank',
             'Amount Due','Amount Paid','Balance Due','Year']
    ws.append(headers)
    for cell in ws[3]:
        cell.font=hf; cell.fill=hfill
        cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=thin
    ws.row_dimensions[3].height=20
    for i,p in enumerate(payments,4):
        row=[safe_str(p.get('receipt_no')),
             p['payment_date'].strftime('%d/%m/%Y') if p.get('payment_date') and hasattr(p['payment_date'],'strftime') else '',
             safe_str(p.get('resident_name')),safe_str(p.get('phone')),
             safe_str(p.get('occupancy_type')),safe_str(p.get('house_no')),
             safe_str(p.get('street')),
             f"{safe_str(p.get('payment_code'))} - {safe_str(p.get('payment_type'))}",
             safe_str(p.get('payment_method')),safe_str(p.get('bank_name')),
             float(p.get('amount_due',0)),float(p.get('amount_paid',0)),
             float(p.get('balance_due',0)),p.get('service_year','')]
        ws.append(row)
        fill=PatternFill("solid",fgColor=light) if i%2==0 else PatternFill("solid",fgColor=white)
        for cell in ws[i]:
            cell.fill=fill; cell.border=thin; cell.font=Font(name="Arial",size=10)
        for col in [11,12,13]: ws.cell(i,col).number_format='#,##0.00'
    tr=ws.max_row+1
    ws.cell(tr,10,"TOTAL")
    for col in [11,12,13]:
        ws.cell(tr,col,f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{tr-1})")
        c=ws.cell(tr,col); c.font=Font(bold=True,name="Arial",color=white)
        c.fill=PatternFill("solid",fgColor=green); c.border=thin; c.number_format='#,##0.00'
    ws.cell(tr,10).font=Font(bold=True,name="Arial",color=white)
    ws.cell(tr,10).fill=PatternFill("solid",fgColor=green); ws.cell(tr,10).border=thin
    for i,w in enumerate([15,12,25,14,12,10,18,22,14,18,14,14,14,8],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


def generate_report_pdf(payments, title, from_date, to_date):
    """Generate a branded PDF payment report"""

    # FPDF uses latin-1 internally — strip/replace any non-latin-1 characters
    def s(text):
        return str(text).replace('\u2014', '-').replace('\u2013', '-')\
                        .replace('\u2019', "'").replace('\u2018', "'")\
                        .replace('\u201c', '"').replace('\u201d', '"')\
                        .encode('latin-1', errors='replace').decode('latin-1')

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Header bar
    pdf.set_fill_color(16,185,129); pdf.rect(0,0,210,38,'F')
    pdf.set_font("Arial","B",16); pdf.set_text_color(255,255,255)
    pdf.cell(0,13,"SUNSHINE ESTATE",ln=True,align="C")
    pdf.set_font("Arial","",10)
    pdf.cell(0,7,"Estate Management System - Payment Report",ln=True,align="C")
    pdf.set_font("Arial","",9)
    pdf.cell(0,7,s(f"Period: {from_date.strftime('%d %b %Y')} to {to_date.strftime('%d %b %Y')}"),ln=True,align="C")
    pdf.set_text_color(0,0,0); pdf.ln(4)

    total_paid    = sum(float(p.get('amount_paid',0)) for p in payments)
    total_balance = sum(float(p.get('balance_due',0)) for p in payments)
    type_totals   = {}
    for p in payments:
        k = f"{p.get('payment_code','?')} - {p.get('payment_type','?')}"
        type_totals[k] = type_totals.get(k,0) + float(p.get('amount_paid',0))

    # Summary bar
    pdf.set_fill_color(240,253,244); pdf.set_font("Arial","B",9)
    pdf.cell(47,8,s(f"Records: {len(payments)}"),border=1,fill=True,align="C")
    pdf.cell(47,8,s(f"Collected: {format_currency_pdf(total_paid)}"),border=1,fill=True,align="C")
    pdf.cell(47,8,s(f"Outstanding: {format_currency_pdf(total_balance)}"),border=1,fill=True,align="C")
    pdf.cell(49,8,s(f"Generated: {datetime.now().strftime('%d/%m/%Y')}"),border=1,fill=True,align="C")
    pdf.ln(10)

    # Breakdown by type
    if type_totals:
        pdf.set_font("Arial","B",9); pdf.set_fill_color(220,252,231)
        pdf.cell(0,6,"  Breakdown by Payment Type",ln=True,fill=True)
        pdf.set_font("Arial","",8)
        for k,v in type_totals.items():
            pdf.cell(130,5,s(f"  {k}"),border='B')
            pdf.cell(60,5,s(format_currency_pdf(v)),border='B',align="R",ln=True)
        pdf.ln(5)

    # Table
    col_w = [30,20,42,20,18,30,20]
    hdrs  = ["Receipt No","Date","Resident","Street","Occ","Payment Type","Paid"]
    pdf.set_fill_color(16,185,129); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",8)
    for i,h in enumerate(hdrs):
        pdf.cell(col_w[i],7,h,border=1,fill=True,align="C")
    pdf.ln(); pdf.set_text_color(0,0,0)

    for idx,p in enumerate(payments):
        pdf.set_fill_color(240,253,244) if idx%2==0 else pdf.set_fill_color(255,255,255)
        pdf.set_font("Arial","",7)
        vals = [
            s(safe_str(p.get('receipt_no','')))[:18],
            p['payment_date'].strftime('%d/%m/%Y') if p.get('payment_date') else '',
            s(safe_str(p.get('resident_name','')))[:24],
            s(safe_str(p.get('street','')))[:12],
            s(safe_str(p.get('occupancy_type','')))[:8],
            s(safe_str(p.get('payment_code','?'))+' '+safe_str(p.get('payment_type','')))[:20],
            s(format_currency_pdf(p.get('amount_paid',0))),
        ]
        for i,val in enumerate(vals):
            pdf.cell(col_w[i],5,val,border=1,fill=True,align="R" if i==6 else "L")
        pdf.ln()

    # Total
    pdf.set_fill_color(16,185,129); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",8)
    pdf.cell(sum(col_w[:6]),7,"TOTAL COLLECTED",border=1,fill=True,align="R")
    pdf.cell(col_w[6],7,s(format_currency_pdf(total_paid)),border=1,fill=True,align="R")
    pdf.ln(10)

    pdf.set_text_color(120,120,120); pdf.set_font("Arial","I",7)
    pdf.cell(0,5,s(f"Sunshine Estate Management System - {datetime.now().strftime('%d %b %Y %H:%M')}"),align="C")

    out = pdf.output(dest='S')
    return bytes(out) if isinstance(out,bytearray) else out.encode('latin-1')


# ── Email Manager ────────────────────────────────────────────────
class EmailManager:
    def __init__(self):
        self.sender=st.session_state.get('email_sender','')
        self.password=st.session_state.get('email_password','')

    def save_config(self,email,password):
        st.session_state['email_sender']=email; st.session_state['email_password']=password
        self.sender=email; self.password=password

    def send(self,recipient,name,receipt_data,pdf_bytes,receipt_no):
        if not self.sender or not self.password: return False,"Email not configured"
        try:
            msg=MIMEMultipart(); msg['From']=self.sender; msg['To']=recipient
            msg['Subject']=f"Payment Receipt - {receipt_no}"
            body=f"Dear {name},\n\nThank you for your payment to Sunshine Estate.\n\nReceipt: {receipt_no}\nAmount Paid: {format_currency(receipt_data['amount_paid'])}\nDate: {receipt_data['payment_date'].strftime('%d/%m/%Y')}\nYear: {receipt_data['service_year']}\n\nBest regards,\nSunshine Estate Management"
            msg.attach(MIMEText(body,'plain'))
            att=MIMEBase('application','pdf'); att.set_payload(pdf_bytes)
            encoders.encode_base64(att)
            att.add_header('Content-Disposition',f'attachment; filename="{receipt_no}.pdf"')
            msg.attach(att)
            with smtplib.SMTP_SSL('smtp.gmail.com',465) as s:
                s.login(self.sender,self.password); s.send_message(msg)
            return True,"✅ Email sent successfully"
        except Exception as e: return False,f"Email error: {str(e)}"


# ════════════════════════════════════════════════════════════════
# MAIN APP
# ════════════════════════════════════════════════════════════════
def main():
    # ── Static Login ─────────────────────────────────────────────
    LOGIN_USERNAME = "admin"
    LOGIN_PASSWORD = "sunshine2024"

    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False

    if not st.session_state.logged_in:
        st.markdown("""
        <style>
            .login-container{max-width:400px;margin:80px auto;padding:40px;
                background:white;border-radius:16px;box-shadow:0 4px 20px rgba(0,0,0,.12)}
            .login-header{background:linear-gradient(90deg,#059669,#10b981);padding:20px;
                border-radius:10px;margin-bottom:28px;color:white;text-align:center}
        </style>
        <div class="login-header"><h1>🏠 Sunshine Estate</h1><p>Management System</p></div>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            st.subheader("🔐 Sign In")
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Login", use_container_width=True)
            if submitted:
                if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
                    st.session_state.logged_in = True
                    st.rerun()
                else:
                    st.error("❌ Invalid username or password.")
        return
    # ─────────────────────────────────────────────────────────────

    st.markdown("""
    <style>
        .main-header{background:linear-gradient(90deg,#059669,#10b981);padding:20px;
            border-radius:10px;margin-bottom:20px;color:white;text-align:center}
        .stat-card{background:white;padding:20px;border-radius:10px;
            box-shadow:0 2px 4px rgba(0,0,0,.1);border-left:4px solid #10b981}
        .payment-history{background:#f0fdf4;padding:15px;border-radius:8px;
            margin:10px 0;border-left:3px solid #10b981}
        .year-badge{background:#10b981;color:white;padding:5px 15px;border-radius:20px;
            font-weight:bold;display:inline-block;margin-bottom:10px}
        .stButton>button{background-color:#10b981;color:white;border:none}
        .stButton>button:hover{background-color:#059669;color:white}
        [data-testid="stSidebar"]{background-color:#f0fdf4}
        .tag-green{background:#10b981;color:white;padding:3px 9px;border-radius:12px;
            font-size:11px;font-weight:bold}
    </style>""", unsafe_allow_html=True)

    estate = EstateManager()
    email_mgr = EmailManager()

    if not estate.conn:
        st.error("⚠️ Database connection failed."); return

    if 'edit_mode' not in st.session_state:
        st.session_state.edit_mode = {}

    # ── Sidebar ──────────────────────────────────────────────────
    st.sidebar.title("🏠 Sunshine Estate")
    st.sidebar.markdown("---")
    if st.sidebar.button("🚪 Logout"):
        st.session_state.logged_in = False
        st.rerun()
    selected = st.sidebar.radio("Navigation", [
        "🏠 Dashboard", "🏘️ Properties", "👥 Residents",
        "💰 Payments", "📊 Reports", "📧 Email Receipts",
        "📤 Upload Data", "⚙️ Settings"
    ])
    st.sidebar.markdown("---")
    st.sidebar.info("💡 Upload your Excel template via Upload Data")

    # ════════════════════════════════════════════════════════════
    # DASHBOARD
    # ════════════════════════════════════════════════════════════
    if selected == "🏠 Dashboard":
        st.markdown('<div class="main-header"><h1>🏠 Estate Dashboard</h1></div>', unsafe_allow_html=True)

        stats = estate.get_stats()
        c1,c2,c3,c4 = st.columns(4)
        for col,label,val in [(c1,"Properties",stats['props']),
                               (c2,"Active Residents",stats['residents']),
                               (c3,"Monthly Collection",format_currency(stats['monthly'])),
                               (c4,"Outstanding Dues",format_currency(stats['outstanding']))]:
            with col:
                st.markdown('<div class="stat-card">',unsafe_allow_html=True)
                st.metric(label,val)
                st.markdown('</div>',unsafe_allow_html=True)

        st.markdown("---")
        st.subheader("🔍 Filter Residents")

        streets = estate.get_streets()
        street_opts = ["All"] + [s['name'] for s in streets]
        prop_types = estate.get_property_types()
        type_opts = ["All"] + [t['name'] for t in prop_types]

        fc1,fc2,fc3,fc4,fc5,fc6 = st.columns(6)
        with fc1: f_status = st.selectbox("Status",["Active Only","All Residents","Inactive Only"])
        with fc2: f_occ = st.selectbox("Occupancy",["All","RESIDENT","LANDLORD","TENANT","UNCOMPLETED","UNDEVELOPED","UNOCCUPIED"])
        with fc3:
            pts_f = estate.get_payment_types()
            pt_opts = ["All"] + [p['code'] for p in pts_f]
            f_pay = st.selectbox("Payment Type", pt_opts)
        with fc4: f_year = st.selectbox("Year",["All"]+list(range(datetime.now().year,datetime.now().year-7,-1)))
        with fc5: f_street = st.selectbox("Street",street_opts)
        with fc6: f_type = st.selectbox("House Type",type_opts)

        filtered = estate.get_filtered_residents(f_pay,f_year,f_occ,f_street,f_status,f_type)

        if filtered:
            active_c = sum(1 for r in filtered if r.get('is_active',1)==1)
            inactive_c = len(filtered)-active_c
            total_c = len(filtered)

            # Sort by last payment date descending (most recent payments first)
            def sort_key(r):
                d = r.get('last_payment_date')
                if d and hasattr(d, 'strftime'): return d
                from datetime import date as _date
                return _date(1900, 1, 1)
            filtered_sorted = sorted(filtered, key=sort_key, reverse=True)

            # Default: show last 10, toggle to show all
            show_all_dash = st.session_state.get('dash_show_all', False)
            display_filtered = filtered_sorted if show_all_dash else filtered_sorted[:10]

            st.info(f"Showing **{len(display_filtered)}** of **{total_c}** resident(s) — 🟢 {active_c} Active, 🔴 {inactive_c} Inactive")

            display=[]
            for r in display_filtered:
                lpd = r.get('last_payment_date')
                lpd_str = lpd.strftime('%d/%m/%Y') if lpd and hasattr(lpd,'strftime') else '—'
                desc_str = safe_str(r.get('last_payment_description') or '—')[:40]
                display.append({
                    'Status':'🟢 Active' if r.get('is_active',1) else '🔴 Inactive',
                    'Name':r['name'],
                    'Property Address':f"{r.get('house_no','?')}, {r.get('street','?')}",
                    'Occupancy':r.get('occupancy_type','N/A'),
                    'Total Due':format_currency(r.get('total_due',0)),
                    'Total Paid':format_currency(r.get('total_paid',0)),
                    'Balance':format_currency(r.get('balance',0)),
                    'Payment Status':r.get('payment_status','N/A'),
                    'Last Payment Date':lpd_str,
                    'Last Payment Description':desc_str,
                })
            df=pd.DataFrame(display)
            st.dataframe(df.astype(str),use_container_width=True,hide_index=True)

            # Show all / show less toggle
            if total_c > 10:
                tog_label = f"📋 Show all {total_c} residents" if not show_all_dash else "📋 Show last 10 only"
                if st.button(tog_label, key="dash_toggle", use_container_width=False):
                    st.session_state['dash_show_all'] = not show_all_dash; st.rerun()

            dl1,dl2=st.columns(2)
            with dl1:
                st.download_button("📥 Download CSV",df.to_csv(index=False),
                    f"residents_{datetime.now().strftime('%Y%m%d')}.csv","text/csv",
                    use_container_width=True)
            with dl2:
                wb2=Workbook(); ws2=wb2.active; ws2.title="Residents"
                ws2.append(list(df.columns))
                for r2 in df.itertuples(index=False): ws2.append(list(r2))
                buf2=io.BytesIO(); wb2.save(buf2)
                st.download_button("📥 Download Excel",buf2.getvalue(),
                    f"residents_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
        else:
            st.info("No residents match the selected filters")

        st.markdown("---")
        if stats['recent']:
            st.subheader("🕐 Last 10 Transactions")
            hc1,hc2,hc3,hc4,hc5,hc6=st.columns([2,2,1,1,2,2])
            with hc1: st.markdown("**Resident**")
            with hc2: st.markdown("**Property Address**")
            with hc3: st.markdown("**Amount**")
            with hc4: st.markdown("**Date**")
            with hc5: st.markdown("**Description**")
            with hc6: st.markdown("**Receipt**")
            st.markdown("<hr style='margin:4px 0'>", unsafe_allow_html=True)
            for p in stats['recent']:
                c1,c2,c3,c4,c5,c6=st.columns([2,2,1,1,2,2])
                prop_addr = f"{p.get('house_no','?')}, {p.get('street','?')}"
                pay_date = p['payment_date'].strftime('%d/%m/%Y') if p.get('payment_date') and hasattr(p['payment_date'],'strftime') else 'N/A'
                desc = str(p.get('payment_description') or p.get('payment_type') or 'N/A')[:30]
                with c1: st.write(f"{p['name']}")
                with c2: st.write(prop_addr)
                with c3: st.write(format_currency(p['amount_paid']))
                with c4: st.write(pay_date)
                with c5: st.write(desc)
                with c6: st.markdown(f'<span class="tag-green">{p.get("receipt_no","N/A")}</span>',unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════
    # PROPERTIES
    # ════════════════════════════════════════════════════════════
    elif selected == "🏘️ Properties":
        st.markdown('<div class="main-header"><h1>🏘️ Property Management</h1></div>', unsafe_allow_html=True)
        tab1,tab2=st.tabs(["📋 View All","➕ Add New"])

        with tab1:
            properties=estate.get_properties()
            if properties:
                streets_dict={}
                for prop in properties:
                    streets_dict.setdefault(prop['street'],[]).append(prop)
                for street,props in sorted(streets_dict.items()):
                    with st.expander(f"📍 {street} ({len(props)} properties)"):
                        for prop in props:
                            resident_display=prop.get('resident_names','') or prop.get('residents','') or 'Vacant'
                            c1,c2,c3=st.columns([2,2,1])
                            with c1:
                                st.write(f"**House {prop['house_no']}**")
                                st.caption(f"Residents: {resident_display}")
                            with c2:
                                st.write(f"{prop.get('type_name',prop.get('type','N/A'))} — {prop['status']}")
                            with c3:
                                sc="🟢" if prop['status']=='Occupied' else "⚪"
                                st.write(sc)
                                if st.session_state.edit_mode.get(f"prop_{prop['id']}",False):
                                    if st.button("❌ Cancel",key=f"cancel_{prop['id']}",use_container_width=True):
                                        st.session_state.edit_mode[f"prop_{prop['id']}"]=False; st.rerun()
                                else:
                                    if st.button("✏️ Edit",key=f"edit_{prop['id']}",use_container_width=True):
                                        st.session_state.edit_mode[f"prop_{prop['id']}"]=True; st.rerun()
                                    if st.button("🗑️ Delete",key=f"delete_{prop['id']}",use_container_width=True):
                                        di=estate.get_property_deletion_info(prop['id'])
                                        if di and not di['can_delete']:
                                            reasons=", ".join(di['reasons'])
                                            st.error(f"❌ Cannot delete. Has: {reasons}")
                                            st.info("💡 Delete all payments and deactivate all residents first.")
                                        else:
                                            ok,msg=estate.delete_property(prop['id'])
                                            if ok: st.success(msg); st.rerun()
                                            else: st.error(msg)

                            if st.session_state.edit_mode.get(f"prop_{prop['id']}",False):
                                with st.form(f"edit_form_{prop['id']}"):
                                    st.subheader("Edit Property")
                                    all_st=estate.get_all_streets()
                                    all_types=estate.get_property_types()
                                    st_map={s['id']:s['name'] for s in all_st}
                                    st_opts=list(st_map.values())
                                    st_ids=list(st_map.keys())
                                    cur_st=next((s['name'] for s in all_st if s['id']==prop['street_id']),st_opts[0] if st_opts else '')
                                    cur_st_idx=st_opts.index(cur_st) if cur_st in st_opts else 0
                                    new_house=st.text_input("House Number",prop['house_no'])
                                    new_street_str=st.selectbox("Street",st_opts,index=cur_st_idx) if st_opts else st.text_input("Street")
                                    new_street_id=st_ids[st_opts.index(new_street_str)] if st_opts else prop['street_id']
                                    type_names=[t['name'] for t in all_types]
                                    cur_type=prop.get('type_name',prop.get('type',''))
                                    cur_type_idx=type_names.index(cur_type) if cur_type in type_names else 0
                                    new_type_str=st.selectbox("Type",type_names,index=cur_type_idx) if type_names else st.text_input("Type")
                                    new_type_id=next((t['id'] for t in all_types if t['name']==new_type_str),prop['type_id']) if all_types else prop['type_id']
                                    status_opts=["Vacant","Occupied","Uncompleted"]
                                    try: si=status_opts.index(prop['status'])
                                    except: si=0
                                    new_status=st.selectbox("Status",status_opts,index=si)
                                    if st.form_submit_button("💾 Save Changes"):
                                        ok,msg=estate.update_property(prop['id'],new_house,new_street_id,new_type_id,new_status)
                                        if ok:
                                            st.success(msg)
                                            st.session_state.edit_mode[f"prop_{prop['id']}"]=False; st.rerun()
                                        else: st.error(msg)
            else:
                st.info("No properties registered yet. Upload data or add manually.")

        with tab2:
            all_streets=estate.get_all_streets()
            all_types=estate.get_property_types()
            if not all_streets: st.warning("⚠️ No streets found. Upload data first.")
            if not all_types: st.warning("⚠️ No property types found. Upload data first.")
            with st.form("add_property_form"):
                st.subheader("Add New Property")
                house_no_in=st.text_input("House Number *")
                st_map={s['id']:s['name'] for s in all_streets}
                st_opts_list=list(st_map.values()); st_ids_list=list(st_map.keys())
                street_sel=st.selectbox("Street *",st_opts_list if st_opts_list else ["No streets"])
                type_names=[t['name'] for t in all_types]
                type_sel=st.selectbox("Type *",type_names if type_names else ["No types"])
                type_id_sel=next((t['id'] for t in all_types if t['name']==type_sel),None)
                street_id_sel=st_ids_list[st_opts_list.index(street_sel)] if st_opts_list and street_sel in st_opts_list else None
                status_in=st.selectbox("Status",["Vacant","Occupied","Uncompleted"])
                if st.form_submit_button("➕ Add Property"):
                    if not house_no_in: st.error("House number is required!")
                    elif not street_id_sel: st.error("Please select a valid street.")
                    elif not type_id_sel: st.error("Please select a valid property type.")
                    else:
                        ok,msg=estate.add_property(house_no_in,street_id_sel,type_id_sel,status_in)
                        if ok: st.success(msg); st.rerun()
                        else: st.error(msg)

    # ════════════════════════════════════════════════════════════
    # RESIDENTS
    # ════════════════════════════════════════════════════════════
    elif selected == "👥 Residents":
        st.markdown('<div class="main-header"><h1>👥 Resident Management</h1></div>', unsafe_allow_html=True)
        tab1,tab2=st.tabs(["📋 View All","➕ Add New"])

        with tab1:
            cf1,cf2=st.columns([1,3])
            with cf1:
                status_filter_res=st.selectbox("Show",["Active Only","All Residents","Inactive Only"],key="res_status_filter")
            with cf2:
                search_filter=st.text_input("🔍 Search residents",placeholder="Name, phone, or street...")

            if status_filter_res=="Active Only": residents=estate.get_residents()
            elif status_filter_res=="Inactive Only": residents=estate.get_residents(include_inactive_only=True)
            else: residents=estate.get_residents(include_all=True)

            if residents:
                if search_filter:
                    residents=[r for r in residents if
                               search_filter.lower() in r['name'].lower() or
                               search_filter.lower() in (r.get('phone') or '').lower() or
                               search_filter.lower() in (r.get('street') or '').lower()]
                active_c=sum(1 for r in residents if r.get('is_active',1)==1)
                inactive_c=len(residents)-active_c
                st.info(f"Showing {len(residents)} resident(s) — 🟢 {active_c} Active, 🔴 {inactive_c} Inactive")

                for resident in residents:
                    bal_color="#ef4444" if float(resident.get('current_balance') or 0)>0 else "#10b981"
                    is_active=resident.get('is_active',1)==1
                    status_badge="🟢 ACTIVE" if is_active else "🔴 INACTIVE"
                    with st.expander(f"{status_badge} | 👤 {resident['name']} — {resident.get('house_no','N/A')}, {resident.get('street','N/A')}"):
                        c1,c2=st.columns([2,1])
                        with c1:
                            st.write(f"**Phone:** {resident.get('phone') or 'Not provided'}")
                            st.write(f"**Email:** {resident.get('email') or 'Not provided'}")
                            st.write(f"**Occupancy:** {resident.get('occupancy_type','N/A')}")
                            st.write(f"**Join Date:** {resident.get('join_date','N/A')}")
                            st.write(f"**Status:** {status_badge}")
                            st.markdown(f"**Current Balance:** <span style='color:{bal_color};font-weight:bold'>{format_currency(resident.get('current_balance',0))}</span>",unsafe_allow_html=True)
                        with c2:
                            if is_active:
                                if st.button("💰 Add Payment",key=f"pay_{resident['id']}"):
                                    st.session_state['selected_resident_for_payment']=resident['id']; st.rerun()
                                if st.button("✏️ Edit",key=f"edit_res_{resident['id']}"):
                                    st.session_state.edit_mode[f"res_{resident['id']}"]=True; st.rerun()
                                current_bal=float(resident.get('current_balance') or 0)
                                if current_bal>0:
                                    st.warning(f"⚠️ Outstanding: {format_currency(current_bal)}")
                                    clear_bal=st.checkbox("Clear balance on deactivate",key=f"clrbal_{resident['id']}",
                                                          help="Check to remove balance record when deactivating")
                                    if st.button("🗑️ Deactivate",key=f"delete_res_{resident['id']}"):
                                        ok,msg=estate.delete_resident(resident['id'],force_delete=clear_bal)
                                        if ok: st.success(msg); st.rerun()
                                        else:
                                            st.error(msg)
                                            if "outstanding balance" in msg.lower():
                                                st.info("💡 Tip: Check 'Clear balance on deactivate' above to remove the balance record")
                                else:
                                    if st.button("🗑️ Deactivate",key=f"delete_res_{resident['id']}"):
                                        ok,msg=estate.delete_resident(resident['id'])
                                        if ok: st.success(msg); st.rerun()
                                        else: st.error(msg)
                            else:
                                if st.button("♻️ Reactivate",key=f"reactivate_{resident['id']}",type="primary"):
                                    ok,msg=estate.reactivate_resident(resident['id'])
                                    if ok: st.success(msg); st.rerun()
                                    else: st.error(msg)

                        if st.session_state.edit_mode.get(f"res_{resident['id']}",False):
                            with st.form(f"edit_resident_form_{resident['id']}"):
                                st.subheader("Edit Resident")
                                new_name=st.text_input("Name",value=resident['name'])
                                new_phone=st.text_input("Phone",value=resident.get('phone',''))
                                new_email=st.text_input("Email",value=resident.get('email',''))
                                all_props=estate.get_properties()
                                prop_opts=[f"{p['house_no']}, {p['street']}" for p in all_props]
                                cur_prop=f"{resident.get('house_no','N/A')}, {resident.get('street','N/A')}"
                                prop_idx=prop_opts.index(cur_prop) if cur_prop in prop_opts else 0
                                new_prop_str=st.selectbox("Property",prop_opts,index=prop_idx)
                                new_prop_id=all_props[prop_opts.index(new_prop_str)]['id'] if prop_opts else resident.get('property_id')
                                occ_opts=['LANDLORD','RESIDENT','TENANT','UNOCCUPIED','UNCOMPLETED','UNDEVELOPED']
                                cur_occ=resident.get('occupancy_type','TENANT')
                                occ_idx=occ_opts.index(cur_occ) if cur_occ in occ_opts else 0
                                new_occ=st.selectbox("Occupancy",occ_opts,index=occ_idx)
                                ca,cb=st.columns(2)
                                with ca:
                                    if st.form_submit_button("💾 Save Changes"):
                                        ok,msg=estate.update_resident(resident['id'],new_name,new_prop_id,new_phone,new_email,new_occ)
                                        if ok:
                                            st.success(msg)
                                            st.session_state.edit_mode[f"res_{resident['id']}"]=False; st.rerun()
                                        else: st.error(msg)
                                with cb:
                                    if st.form_submit_button("❌ Cancel"):
                                        st.session_state.edit_mode[f"res_{resident['id']}"]=False; st.rerun()

                        # Payment History by Year (original format)
                        st.markdown("---")
                        st.subheader("📊 Payment History by Year")
                        payment_history=estate.get_resident_payment_history(resident['id'])
                        if payment_history:
                            for yr in sorted(payment_history.keys(),reverse=True):
                                info=payment_history[yr]
                                st.markdown('<div class="payment-history">',unsafe_allow_html=True)
                                st.markdown(f'<span class="year-badge">📅 {yr}</span>',unsafe_allow_html=True)
                                ya,yb,yc,yd=st.columns(4)
                                with ya: st.metric("Payments",f"{info['payment_count']}")
                                with yb: st.metric("Total Paid",format_currency(info['total_paid']))
                                with yc: st.metric("Total Due",format_currency(info['total_due']))
                                with yd:
                                    bv=info['balance']
                                    bi="🔴" if bv>0 else "🟢"
                                    st.metric(f"{bi} Balance",format_currency(bv))
                                if info['outstanding_brought_forward']>0:
                                    st.caption(f"⬆️ Outstanding brought forward: {format_currency(info['outstanding_brought_forward'])}")
                                if info['last_payment_date']:
                                    st.caption(f"Last payment: {info['last_payment_date'].strftime('%d/%m/%Y')}")
                                # Individual payment receipts
                                pmts=estate.get_resident_payments(resident['id'])
                                yr_pmts=[p for p in pmts if p.get('service_year')==yr]
                                for p in yr_pmts:
                                    pc1,pc2,pc3=st.columns([3,1,1])
                                    with pc1:
                                        st.markdown(f'<span class="tag-green">{p["receipt_no"]}</span> {p.get("code","?")} — {p.get("payment_type","N/A")}',unsafe_allow_html=True)
                                    with pc2: st.text(format_currency(p['amount_paid']))
                                    with pc3:
                                        rd=estate.get_receipt(p['receipt_no'])
                                        if rd:
                                            pdf_b=generate_pdf(rd)
                                            st.download_button("📥",pdf_b,f"{p['receipt_no']}.pdf",
                                                "application/pdf",key=f"dl_{p['id']}")
                                st.markdown('</div>',unsafe_allow_html=True)
                            # Carry forward option
                            st.markdown("---")
                            latest_yr=max(payment_history.keys())
                            if payment_history[latest_yr]['balance']>0:
                                cfa,cfb=st.columns([3,1])
                                with cfa:
                                    st.info(f"Outstanding for {latest_yr}: {format_currency(payment_history[latest_yr]['balance'])}")
                                with cfb:
                                    if st.button(f"Carry to {latest_yr+1}",key=f"cf_{resident['id']}"):
                                        ok,msg=estate.carry_forward_outstanding(resident['id'],latest_yr,latest_yr+1)
                                        if ok: st.success(msg); st.rerun()
                                        else: st.warning(msg)
                        else:
                            st.info("No payment history available")

            else:
                if status_filter_res=="Inactive Only":
                    st.info("No inactive residents found. All residents are active!")
                elif status_filter_res=="All Residents":
                    st.warning("No residents in the database. Add your first resident or upload data.")
                else:
                    st.warning("No active residents. Try 'All Residents' to see inactive ones.")

        with tab2:
            props=estate.get_properties()
            if not props: st.warning("⚠️ No properties found. Upload data or add properties first.")
            with st.form("add_resident_form"):
                st.subheader("Add New Resident")
                c1,c2=st.columns(2)
                with c1:
                    name_in=st.text_input("Full Name *")
                    phone_in=st.text_input("Phone *",placeholder="08012345678")
                    email_in=st.text_input("Email (Optional)")
                with c2:
                    occ_types=['LANDLORD','RESIDENT','TENANT','UNOCCUPIED','UNCOMPLETED','UNDEVELOPED']
                    occ_in=st.selectbox("Occupancy Type *",occ_types,index=2)
                    join_in=st.date_input("Join Date",date.today())
                    prop_opts=[f"{p['house_no']} — {p['street']}" for p in props] if props else ["No properties available"]
                    prop_sel=st.selectbox("Select Property *",prop_opts)
                if st.form_submit_button("➕ Add Resident"):
                    if not name_in or not phone_in:
                        st.error("Name and phone are required!")
                    elif not props or prop_sel=="No properties available":
                        st.error("Please add properties first!")
                    else:
                        prop_id=next((p['id'] for p in props if f"{p['house_no']} — {p['street']}"==prop_sel),None)
                        if not prop_id:
                            st.error("Selected property not found. Please refresh.")
                        else:
                            ok,rid,msg=estate.add_resident(name_in,prop_id,phone_in,email_in,occ_in,join_in)
                            if ok: st.success(f"✅ {msg}"); st.balloons(); st.rerun()
                            else: st.error(msg)

    # ════════════════════════════════════════════════════════════
    # PAYMENTS
    # ════════════════════════════════════════════════════════════
    elif selected == "💰 Payments":
        st.markdown('<div class="main-header"><h1>💰 Payment Management</h1></div>', unsafe_allow_html=True)

        # Handle redirect from Residents page
        if 'selected_resident_for_payment' in st.session_state:
            st.info("💡 Scroll to 'Record Payment' tab — resident pre-selected from Residents page")

        tab1,tab2=st.tabs(["➕ Record Payment","📋 View All Payments"])

        with tab1:
            search=st.text_input("🔍 Search Resident",placeholder="Enter name or phone...")
            if search:
                residents=estate.search_residents(search)
                if residents:
                    sel_r=st.selectbox("Select Resident",residents,
                        format_func=lambda r: f"{r['name']} — {r.get('phone','N/A')} — {r.get('house_no','?')}, {r.get('street','?')}")
                    if sel_r:
                        bal=float(sel_r.get('current_balance') or 0)
                        bc="#ef4444" if bal>0 else "#10b981"
                        st.markdown(f"Current Balance: <span style='color:{bc};font-weight:bold'>{format_currency(bal)}</span>",unsafe_allow_html=True)
                        if sel_r.get('occupancy_type')=='LANDLORD':
                            st.info("🏠 Landlord — annual service charges are ₦0")
                        # Show last saved receipt download OUTSIDE the form
                        if st.session_state.get('last_receipt_pdf') and st.session_state.get('last_receipt_no'):
                            st.success(f"✅ Payment recorded — Receipt: **{st.session_state['last_receipt_no']}**")
                            col_dl, col_dis = st.columns([2,1])
                            with col_dl:
                                st.download_button("📥 Download Receipt", st.session_state['last_receipt_pdf'],
                                    f"{st.session_state['last_receipt_no']}.pdf", "application/pdf", key="dl_new_receipt")
                            with col_dis:
                                if st.button("✖️ Dismiss", key="dismiss_receipt"):
                                    del st.session_state['last_receipt_pdf']
                                    del st.session_state['last_receipt_no']
                                    st.rerun()

                        with st.form("payment_form"):
                            c1,c2=st.columns(2)
                            with c1:
                                amount_in=st.number_input("Amount Paid *",min_value=0.0,step=1000.0)
                                pay_date=st.date_input("Payment Date *",value=date.today())
                                svc_year=st.number_input("Service Year *",min_value=2015,max_value=2035,value=datetime.now().year)
                                pay_types=estate.get_payment_types()
                                if pay_types:
                                    pt_id=st.selectbox("Payment Type *",[pt['id'] for pt in pay_types],
                                        format_func=lambda x: next((f"{p['code']} — {p['name']} (₦{float(p.get('default_amount',0)):,.0f})" for p in pay_types if p['id']==x),''))
                                else:
                                    st.warning("No payment types found. Add them in Settings first."); pt_id=None
                            with c2:
                                methods=estate.get_payment_methods()
                                pm_id=st.selectbox("Payment Method",[m['id'] for m in methods] if methods else [None],
                                    format_func=lambda x: next((m['name'] for m in methods if m['id']==x),'N/A')) if methods else None
                                banks=estate.get_bank_accounts()
                                ba_id=st.selectbox("Bank Account",[b['id'] for b in banks] if banks else [None],
                                    format_func=lambda x: next((f"{b['bank_name']} — {b['account_number']}" for b in banks if b['id']==x),'N/A')) if banks else None
                                desc_in=st.text_area("Payment Description",placeholder="Payment notes...")
                            submitted=st.form_submit_button("💾 Record Payment",use_container_width=True,type="primary")
                            if submitted:
                                if amount_in>0 and pt_id:
                                    ok,receipt_no,msg=estate.add_payment(
                                        sel_r['id'],sel_r['property_id'],amount_in,pay_date,
                                        svc_year,pt_id,pm_id,ba_id,desc_in)
                                    if ok:
                                        rd=estate.get_receipt(receipt_no)
                                        if rd:
                                            st.session_state['last_receipt_pdf']=generate_pdf(rd)
                                            st.session_state['last_receipt_no']=receipt_no
                                        st.balloons(); st.rerun()
                                    else: st.error(msg)
                                else: st.error("Amount and payment type are required")
                else:
                    st.warning("No residents found")

        with tab2:
            st.subheader("All Payments")
            # Payment filters
            pf1,pf2,pf3,pf4=st.columns(4)
            with pf1:
                streets_pv=estate.get_streets()
                f_st_pv=st.selectbox("Street",["All"]+[s['name'] for s in streets_pv],key="pv_st")
            with pf2:
                pts_pv=estate.get_payment_types()
                f_pt_pv=st.selectbox("Payment Type",["All"]+[f"{p['code']} — {p['name']}" for p in pts_pv],key="pv_pt")
            with pf3:
                f_yr_pv=st.selectbox("Year",["All"]+list(range(datetime.now().year,datetime.now().year-7,-1)),key="pv_yr")
            with pf4:
                srch_pv=st.text_input("🔍 Search name/receipt",key="pv_srch")

            cur_pv=estate.conn.cursor(dictionary=True)
            try:
                cur_pv.execute("""SELECT p.*, r.name as resident_name, r.phone, r.occupancy_type,
                                         pr.house_no, s.name as street,
                                         pt.code as payment_code, pt.name as payment_type,
                                         pm.name as payment_method,
                                         ba.bank_name, ba.account_number
                                  FROM payments p
                                  JOIN residents r ON p.resident_id=r.id
                                  LEFT JOIN properties pr ON p.property_id=pr.id
                                  LEFT JOIN streets s ON pr.street_id=s.id
                                  LEFT JOIN payment_types pt ON p.payment_type_id=pt.id
                                  LEFT JOIN payment_methods pm ON p.payment_method_id=pm.id
                                  LEFT JOIN bank_accounts ba ON p.bank_account_id=ba.id
                                  ORDER BY p.payment_date DESC, p.id DESC""")
                all_pmts=cur_pv.fetchall()
            finally: cur_pv.close()

            if f_st_pv!="All": all_pmts=[p for p in all_pmts if p.get('street')==f_st_pv]
            if f_pt_pv!="All": all_pmts=[p for p in all_pmts if safe_str(p.get('payment_code'))==f_pt_pv.split(' — ')[0]]
            if f_yr_pv!="All": all_pmts=[p for p in all_pmts if str(p.get('service_year'))==str(f_yr_pv)]
            if srch_pv: all_pmts=[p for p in all_pmts if srch_pv.lower() in p.get('resident_name','').lower() or srch_pv.lower() in p.get('receipt_no','').lower()]

            st.info(f"Showing **{len(all_pmts)}** payment(s)")
            for pmt in all_pmts:
                with st.expander(f"🧾 {pmt['receipt_no']} — {pmt.get('resident_name','N/A')} — {format_currency(pmt['amount_paid'])}"):
                    ec1,ec2,ec3=st.columns([3,1,1])
                    with ec1:
                        st.write(f"**Resident:** {pmt.get('resident_name','N/A')} | **Phone:** {pmt.get('phone','N/A')}")
                        st.write(f"**Property Address:** {pmt.get('house_no','?')}, {pmt.get('street','?')} | **Occupancy:** {pmt.get('occupancy_type','N/A')}")
                        st.write(f"**Type:** {pmt.get('payment_code','?')} — {pmt.get('payment_type','N/A')} | **Method:** {pmt.get('payment_method','N/A')}")
                        if pmt.get('bank_name'): st.write(f"**Bank:** {pmt['bank_name']} — {pmt.get('account_number','')}")
                        st.write(f"**Amount:** {format_currency(pmt['amount_paid'])} | **Balance Due:** {format_currency(pmt.get('balance_due',0))} | **Year:** {pmt.get('service_year','N/A')}")
                        if pmt.get('payment_description'): st.write(f"**Note:** {pmt['payment_description']}")
                    with ec2:
                        rd=estate.get_receipt(pmt['receipt_no'])
                        if rd:
                            pdf_b=generate_pdf(rd)
                            st.download_button("📥 Receipt",pdf_b,f"{pmt['receipt_no']}.pdf",
                                "application/pdf",key=f"dl_pv_{pmt['id']}",use_container_width=True)
                        if st.button("✏️ Edit",key=f"edit_pmt_{pmt['id']}",use_container_width=True):
                            st.session_state[f"edit_pmt_{pmt['id']}"]=True
                    with ec3:
                        if st.button("🗑️ Delete",key=f"del_pmt_{pmt['id']}",use_container_width=True):
                            ok,msg=estate.delete_payment(pmt['id'])
                            if ok: st.success(msg); st.rerun()
                            else: st.error(msg)
                    if st.session_state.get(f"edit_pmt_{pmt['id']}"):
                        with st.form(f"form_edit_pmt_{pmt['id']}"):
                            pf1e,pf2e=st.columns(2)
                            with pf1e:
                                new_amt=st.number_input("Amount Paid",value=float(pmt['amount_paid']),min_value=0.0)
                                new_date=st.date_input("Payment Date",value=pmt['payment_date'] if pmt.get('payment_date') else date.today())
                                new_year=st.number_input("Service Year",value=int(pmt['service_year']),min_value=2015,max_value=2035)
                            with pf2e:
                                pts_e=estate.get_payment_types()
                                new_pt=st.selectbox("Type",[p['id'] for p in pts_e],
                                    format_func=lambda x: next((f"{p['code']} — {p['name']}" for p in pts_e if p['id']==x),''),
                                    index=next((i for i,p in enumerate(pts_e) if p['id']==pmt.get('payment_type_id')),0))
                                meths_e=estate.get_payment_methods()
                                new_pm=st.selectbox("Method",[m['id'] for m in meths_e] if meths_e else [None],
                                    format_func=lambda x: next((m['name'] for m in meths_e if m['id']==x),'')) if meths_e else None
                                banks_e=estate.get_bank_accounts()
                                new_ba=st.selectbox("Bank",[b['id'] for b in banks_e] if banks_e else [None],
                                    format_func=lambda x: next((f"{b['bank_name']} — {b['account_number']}" for b in banks_e if b['id']==x),'')) if banks_e else None
                                new_desc=st.text_area("Description",value=safe_str(pmt.get('payment_description')))
                            sb1, sb2 = st.columns(2)
                            with sb1:
                                if st.form_submit_button("💾 Save Changes", use_container_width=True):
                                    ok,msg=estate.update_payment(pmt['id'],new_amt,new_date,new_pt,new_pm,new_ba,new_desc,new_year)
                                    if ok:
                                        st.success(msg); del st.session_state[f"edit_pmt_{pmt['id']}"]; st.rerun()
                                    else: st.error(msg)
                            with sb2:
                                if st.form_submit_button("❌ Cancel", use_container_width=True):
                                    del st.session_state[f"edit_pmt_{pmt['id']}"]; st.rerun()

    # ════════════════════════════════════════════════════════════
    # REPORTS
    # ════════════════════════════════════════════════════════════
    elif selected == "📊 Reports":
        st.markdown('<div class="main-header"><h1>📊 Reports & Analytics</h1></div>', unsafe_allow_html=True)

        now = datetime.now()
        yr  = now.year

        # ── Step 1: Date Range ────────────────────────────────
        st.subheader("🗓️ Step 1 — Select Date Range")
        rc1, rc2 = st.columns(2)
        with rc1:
            from_date = st.date_input("📅 Date From", value=date(yr, 1, 1), key="rd_from")
        with rc2:
            to_date   = st.date_input("📅 Date To",   value=date.today(),   key="rd_to")

        if from_date > to_date:
            st.error("⚠️ 'Date From' cannot be after 'Date To'. Please fix the range.")
        else:
            delta = (to_date - from_date).days + 1
            st.caption(f"📆 **{delta} day(s)** — {from_date.strftime('%d %b %Y')} → {to_date.strftime('%d %b %Y')}")

            # ── Step 2: Filters shown after date range is set ────
            st.markdown("---")
            st.subheader("🔎 Step 2 — Apply Filters (Optional)")
            st.caption("All filters are optional. Leave blank / 'All' to include everything.")

            streets_list = ['All'] + [s['name'] for s in estate.get_streets()]
            pt_list      = estate.get_payment_types()
            pt_opts      = ['All'] + [f"{p['code']} — {p['name']}" for p in pt_list]
            ht_list      = ['All'] + [h['name'] for h in estate.get_property_types()]
            occ_opts     = ['All','LANDLORD','RESIDENT','TENANT','UNOCCUPIED','UNCOMPLETED','UNDEVELOPED']

            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                r_resident_name = st.text_input("👤 Resident's Name",   placeholder="Search by name…",  key="rf_name")
                r_street        = st.selectbox("🏘️ Street Name",         streets_list,                   key="rf_street")
                r_house_no      = st.text_input("🏠 House Number",        placeholder="e.g. 12B",         key="rf_houseno")
            with fc2:
                r_house_type    = st.selectbox("🏗️ House Type",           ht_list,                        key="rf_htype")
                r_occ           = st.selectbox("👥 Occupancy",             occ_opts,                       key="rf_occ")
                r_phone         = st.text_input("📞 Phone No.",            placeholder="Search by phone…", key="rf_phone")
            with fc3:
                r_email         = st.text_input("📧 Email",                placeholder="Search by email…", key="rf_email")
                r_pt_sel        = st.selectbox("💳 Service Payment Type",  pt_opts,                        key="rf_pt")
                amc1, amc2      = st.columns(2)
                with amc1:
                    r_amt_min   = st.number_input("₦ Amount Min", min_value=0.0, value=0.0, step=1000.0, key="rf_amin")
                with amc2:
                    r_amt_max   = st.number_input("₦ Amount Max", min_value=0.0, value=0.0, step=1000.0,
                                                   help="Leave 0 for no upper limit", key="rf_amax")

            st.markdown("---")
            if st.button("🔍 Generate Report", use_container_width=True, type="primary"):
                pt_code     = r_pt_sel.split(' — ')[0] if r_pt_sel != "All" else "All"
                amt_min_val = float(r_amt_min) if r_amt_min > 0 else None
                amt_max_val = float(r_amt_max) if r_amt_max > 0 else None
                with st.spinner("Generating report..."):
                    pmts = estate.get_payments_by_date_range(
                        from_date, to_date,
                        occupancy      = r_occ,
                        payment_type   = pt_code,
                        street         = r_street,
                        year           = 'All',
                        resident_name  = r_resident_name.strip(),
                        house_no       = r_house_no.strip(),
                        house_type     = r_house_type,
                        phone          = r_phone.strip(),
                        email          = r_email.strip(),
                        amount_min     = amt_min_val,
                        amount_max     = amt_max_val,
                    )
                st.session_state['rpt_results'] = pmts
                st.session_state['rpt_label']   = f"{from_date.strftime('%d %b %Y')} to {to_date.strftime('%d %b %Y')}"

            pmts      = st.session_state.get('rpt_results', None)
            rpt_label = st.session_state.get('rpt_label', '')

            if pmts is not None:
                if pmts:
                    total_paid    = sum(float(p.get('amount_paid',0)) for p in pmts)
                    total_due     = sum(float(p.get('amount_due',0))   for p in pmts)
                    total_balance = sum(float(p.get('balance_due',0))  for p in pmts)

                    st.success(f"✅ **{len(pmts)}** payment(s) — {from_date.strftime('%d %b %Y')} to {to_date.strftime('%d %b %Y')}")
                    sm1,sm2,sm3 = st.columns(3)
                    with sm1: st.metric("💰 Total Collected", format_currency(total_paid))
                    with sm2: st.metric("📋 Total Due",        format_currency(total_due))
                    with sm3: st.metric("⚠️ Outstanding",      format_currency(total_balance))

                    type_summary = {}
                    for p in pmts:
                        k = f"{p.get('payment_code','?')} — {p.get('payment_type','N/A')}"
                        type_summary[k] = type_summary.get(k,0)+float(p.get('amount_paid',0))
                    if type_summary:
                        st.markdown("**💳 By Payment Type:**")
                        tc = st.columns(min(len(type_summary),4))
                        for i,(k,v) in enumerate(type_summary.items()):
                            with tc[i%len(tc)]: st.metric(k, format_currency(v))

                    street_sum = {}
                    for p in pmts:
                        k = p.get('street','Unknown')
                        street_sum[k] = street_sum.get(k,0)+float(p.get('amount_paid',0))
                    if len(street_sum) > 1:
                        with st.expander("📍 Breakdown by Street"):
                            for sn,av in sorted(street_sum.items(), key=lambda x:-x[1]):
                                sa,sb = st.columns([3,1])
                                with sa: st.write(f"**{sn}**")
                                with sb: st.write(format_currency(av))

                    st.markdown("---")
                    st.subheader("📋 Full Payment List")
                    df_r = pd.DataFrame([{
                        'Receipt No':   p.get('receipt_no',''),
                        'Payment Date': p['payment_date'].strftime('%d/%m/%Y') if p.get('payment_date') else '',
                        'Resident':     p.get('resident_name',''),
                        'Phone':        p.get('phone',''),
                        'Email':        p.get('email',''),
                        'Occupancy':    p.get('occupancy_type',''),
                        'House No':     p.get('house_no',''),
                        'House Type':   p.get('house_type',''),
                        'Street':       p.get('street',''),
                        'Payment Type': f"{p.get('payment_code','?')} — {p.get('payment_type','N/A')}",
                        'Method':       p.get('payment_method',''),
                        'Bank':         p.get('bank_name',''),
                        'Amount Due':   float(p.get('amount_due',0)),
                        'Amount Paid':  float(p.get('amount_paid',0)),
                        'Balance Due':  float(p.get('balance_due',0)),
                        'Service Year': p.get('service_year',''),
                    } for p in pmts])
                    st.dataframe(df_r.astype(str), use_container_width=True, hide_index=True)

                    rpt_title = f"Sunshine Estate — {from_date.strftime('%d %b %Y')} to {to_date.strftime('%d %b %Y')}"
                    fn_base   = f"report_{from_date}_{to_date}"
                    dl1,dl2,dl3 = st.columns(3)
                    with dl1:
                        xlsx = generate_report_excel(pmts, rpt_title)
                        st.download_button("📥 Download Excel", xlsx, f"{fn_base}.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, type="primary")
                    with dl2:
                        st.download_button("📥 Download CSV", df_r.to_csv(index=False),
                            f"{fn_base}.csv", "text/csv", use_container_width=True)
                    with dl3:
                        pdf_rep = generate_report_pdf(pmts, rpt_title, from_date, to_date)
                        st.download_button("📥 Download PDF", pdf_rep, f"{fn_base}.pdf",
                            "application/pdf", use_container_width=True)
                else:
                    st.info(f"No payments found for **{from_date.strftime('%d %b %Y')}** to **{to_date.strftime('%d %b %Y')}** with the chosen filters.")
                    st.caption("Try widening the date range or changing the filters above.")

        st.markdown("---")
        st.subheader("📈 Annual Payment Summary by Type")
        sy = st.selectbox("Select Year", ["All"]+list(range(yr,yr-10,-1)), key="sum_yr")
        summary = estate.get_payment_summary_by_type(int(sy) if sy != "All" else None)
        if summary:
            grand = sum(float(s.get('total_paid',0)) for s in summary)
            st.metric("Grand Total Collected", format_currency(grand))
            st.markdown("---")
            for s in summary:
                sc1,sc2,sc3,sc4 = st.columns([3,1,1,1])
                with sc1: st.write(f"**{s.get('code','?')} — {s.get('name','N/A')}**")
                with sc2: st.write(f"{s.get('cnt',0)} payment(s)")
                with sc3: st.write(format_currency(s.get('total_paid',0)))
                pct = float(s.get('total_paid',0))/grand*100 if grand>0 else 0
                with sc4: st.write(f"{pct:.1f}%")
        else:
            st.info("No payment data available yet")

    # ════════════════════════════════════════════════════════════
    # EMAIL RECEIPTS
    # ════════════════════════════════════════════════════════════
    elif selected == "📧 Email Receipts":
        st.markdown('<div class="main-header"><h1>📧 Email Receipt Manager</h1></div>', unsafe_allow_html=True)
        tab1,tab2=st.tabs(["🔍 Search & Send","⚙️ Email Settings"])

        with tab1:
            term=st.text_input("Enter name, phone, or email",placeholder="Search...")
            if term:
                residents=estate.search_residents(term)
                if residents:
                    st.success(f"Found **{len(residents)}** resident(s)")
                    for r in residents:
                        bc="#ef4444" if float(r.get('balance') or 0)>0 else "#10b981"
                        with st.expander(f"👤 {r['name']} — {r.get('phone','N/A')}"):
                            ec1,ec2=st.columns([2,1])
                            with ec1:
                                st.write(f"**Email:** {r.get('email','❌ No email on file')}")
                                st.write(f"**Property Address:** {r.get('house_no','?')}, {r.get('street','?')}")
                                st.markdown(f"**Balance:** <span style='color:{bc};font-weight:bold'>{format_currency(r.get('balance',0))}</span>",unsafe_allow_html=True)
                            pmts=estate.get_resident_payments(r['id'])
                            if pmts:
                                st.markdown("---"); st.write("**Payment History:**")
                                for p in pmts:
                                    pa,pb,pc=st.columns([2,1,1])
                                    with pa:
                                        st.markdown(f'<span class="tag-green">{p["receipt_no"]}</span> {p.get("code","?")} ({p["service_year"]})',unsafe_allow_html=True)
                                    with pb: st.text(f"{format_currency(p['amount_paid'])} — {p['payment_date'].strftime('%d/%m/%Y') if p.get('payment_date') else ''}")
                                    with pc:
                                        if r.get('email'):
                                            if st.button("📧 Send",key=f"send_{p['receipt_no']}"):
                                                rd=estate.get_receipt(p['receipt_no'])
                                                if rd:
                                                    pdf_b=generate_pdf(rd)
                                                    ok,msg=email_mgr.send(r['email'],r['name'],rd,pdf_b,p['receipt_no'])
                                                    if ok: st.success(msg)
                                                    else: st.error(msg)
                                        else: st.caption("No email")
                else: st.warning("No residents found")

        with tab2:
            st.info("💡 Use Gmail App Password for authentication.")
            with st.form("email_config_form"):
                c1,c2=st.columns(2)
                with c1: email_in=st.text_input("Sender Email Address",email_mgr.sender,placeholder="your-email@gmail.com")
                with c2: pass_in=st.text_input("Gmail App Password",type="password",placeholder="16-character app password")
                st.markdown("**How to get Gmail App Password:**\n1. Google Account → Security → 2-Step Verification\n2. App passwords → Generate → Copy here")
                if st.form_submit_button("💾 Save Configuration",use_container_width=True):
                    if email_in and pass_in:
                        email_mgr.save_config(email_in,pass_in); st.success("✅ Email configuration saved!")
                    else: st.error("Both email and password are required!")

    # ════════════════════════════════════════════════════════════
    # UPLOAD DATA
    # ════════════════════════════════════════════════════════════
    elif selected == "📤 Upload Data":
        st.markdown('<div class="main-header"><h1>📤 Upload Data</h1></div>', unsafe_allow_html=True)
        st.markdown("""
        <div style="background:#f0fdf4;border-left:5px solid #10b981;padding:18px;border-radius:8px;margin-bottom:15px">
        <h4 style="margin:0 0 10px 0;color:#065f46">📋 How it works — Excel text is extracted into database tables</h4>
        <p style="margin:0 0 10px 0;color:#374151">The system <strong>does not store the Excel file</strong>. It reads the file, extracts the text from each column, and saves it directly into the matching database table:</p>
        <table style="width:100%;border-collapse:collapse;font-size:13px;margin:8px 0;border:1px solid #d1fae5">
        <tr style="background:#065f46;color:white">
          <th style="padding:7px 10px;text-align:left">Excel Column (exact name)</th>
          <th style="padding:7px 10px;text-align:center">→</th>
          <th style="padding:7px 10px;text-align:left">Database Table</th>
          <th style="padding:7px 10px;text-align:left">Saved into column</th>
        </tr>
        <tr style="background:#f0fdf4">
          <td style="padding:6px 10px;font-family:monospace;color:#065f46">STREET</td>
          <td style="padding:6px 10px;text-align:center;color:#10b981;font-weight:bold">→</td>
          <td style="padding:6px 10px;font-weight:bold">streets</td>
          <td style="padding:6px 10px;color:#6b7280">name &nbsp;(one row per unique street)</td>
        </tr>
        <tr>
          <td style="padding:6px 10px;font-family:monospace;color:#065f46">HOUSE NO. &nbsp;+&nbsp; HOUSE TYPE</td>
          <td style="padding:6px 10px;text-align:center;color:#10b981;font-weight:bold">→</td>
          <td style="padding:6px 10px;font-weight:bold">properties</td>
          <td style="padding:6px 10px;color:#6b7280">house_no, type_id, street_id</td>
        </tr>
        <tr style="background:#f0fdf4">
          <td style="padding:6px 10px;font-family:monospace;color:#065f46">RESIDENT'S NAME &nbsp;+&nbsp; PHONE NO. &nbsp;+&nbsp; EMAIL &nbsp;+&nbsp; OCCUPANCY</td>
          <td style="padding:6px 10px;text-align:center;color:#10b981;font-weight:bold">→</td>
          <td style="padding:6px 10px;font-weight:bold">residents</td>
          <td style="padding:6px 10px;color:#6b7280">name, phone, email, occupancy_type</td>
        </tr>
        <tr>
          <td style="padding:6px 10px;font-family:monospace;color:#065f46">PAYMENT &nbsp;(service charge paid)</td>
          <td style="padding:6px 10px;text-align:center;color:#10b981;font-weight:bold">→</td>
          <td style="padding:6px 10px;font-weight:bold">payments</td>
          <td style="padding:6px 10px;color:#6b7280">amount_paid &nbsp;(type = SVC)</td>
        </tr>
        <tr style="background:#f0fdf4">
          <td style="padding:6px 10px;font-family:monospace;color:#065f46">INFRA.</td>
          <td style="padding:6px 10px;text-align:center;color:#10b981;font-weight:bold">→</td>
          <td style="padding:6px 10px;font-weight:bold">payments</td>
          <td style="padding:6px 10px;color:#6b7280">amount_paid &nbsp;(type = INFRA)</td>
        </tr>
        <tr>
          <td style="padding:6px 10px;font-family:monospace;color:#065f46">LEGAL</td>
          <td style="padding:6px 10px;text-align:center;color:#10b981;font-weight:bold">→</td>
          <td style="padding:6px 10px;font-weight:bold">payments</td>
          <td style="padding:6px 10px;color:#6b7280">amount_paid &nbsp;(type = LEGAL)</td>
        </tr>
        <tr style="background:#f0fdf4">
          <td style="padding:6px 10px;font-family:monospace;color:#065f46">TRANSF.</td>
          <td style="padding:6px 10px;text-align:center;color:#10b981;font-weight:bold">→</td>
          <td style="padding:6px 10px;font-weight:bold">payments</td>
          <td style="padding:6px 10px;color:#6b7280">amount_paid &nbsp;(type = TRANSF)</td>
        </tr>
        <tr>
          <td style="padding:6px 10px;font-family:monospace;color:#065f46">DEV. LEVY</td>
          <td style="padding:6px 10px;text-align:center;color:#10b981;font-weight:bold">→</td>
          <td style="padding:6px 10px;font-weight:bold">payments</td>
          <td style="padding:6px 10px;color:#6b7280">amount_paid &nbsp;(type = DEV)</td>
        </tr>
        <tr style="background:#f0fdf4">
          <td style="padding:6px 10px;font-family:monospace;color:#065f46">YR. END PARTY</td>
          <td style="padding:6px 10px;text-align:center;color:#10b981;font-weight:bold">→</td>
          <td style="padding:6px 10px;font-weight:bold">payments</td>
          <td style="padding:6px 10px;color:#6b7280">amount_paid &nbsp;(type = PARTY)</td>
        </tr>
        <tr>
          <td style="padding:6px 10px;font-family:monospace;color:#065f46">BFWD rows &nbsp;(SVC STATUS / SERVICE CHARGE value)</td>
          <td style="padding:6px 10px;text-align:center;color:#10b981;font-weight:bold">→</td>
          <td style="padding:6px 10px;font-weight:bold">balances</td>
          <td style="padding:6px 10px;color:#6b7280">total_due, balance &nbsp;(opening balance per year)</td>
        </tr>
        </table>
        <p style="font-size:12px;color:#065f46;margin:8px 0 0 0">
          ✅ The Excel file is <strong>not stored</strong> — only the text data inside it is extracted &nbsp;|&nbsp;
          ✅ Both sheets supported: <strong>Residents Record</strong> and <strong>Data Template</strong> &nbsp;|&nbsp;
          ✅ Landlords → annual charges set to ₦0 automatically
        </p>
        </div>
        """, unsafe_allow_html=True)

        col_up,col_info=st.columns([2,1])
        with col_up:
            uploaded=st.file_uploader("Choose Excel file (.xlsx or .xls)",type=['xlsx','xls'])
            if uploaded:
                st.success(f"✅ {uploaded.name} ({uploaded.size:,} bytes)")
                try:
                    if uploaded.name.lower().endswith('.xls'):
                        try:
                            all_sheets=pd.read_excel(uploaded,sheet_name=None,engine='xlrd')
                        except Exception as xls_err:
                            st.error(f"Could not read .xls file: {xls_err}. Please save as .xlsx and re-upload.")
                            all_sheets={}
                    else:
                        uploaded.seek(0)
                        all_sheets=pd.read_excel(uploaded,sheet_name=None,engine='openpyxl')

                    if all_sheets:
                        sheet_names=list(all_sheets.keys())
                        st.info(f"📑 Found **{len(sheet_names)}** sheet(s): {', '.join(sheet_names)}")
                        sel_sheet=st.selectbox("Select sheet to upload",sheet_names)
                        preview_df=all_sheets[sel_sheet]
                        st.write(f"**Preview ({len(preview_df)} rows):**")
                        st.dataframe(preview_df.head(10).astype(str), width="stretch")

                        # Pre-upload extraction preview (uses a copy — does NOT modify the original df)
                        with st.expander("🔍 Preview what will be extracted from this file before saving"):
                            st.caption("See what data the system will read and where it will go — before anything is saved to the database.")
                            prev_df_copy = preview_df.copy()
                            prev_cols_upper = [str(c).strip().upper() for c in prev_df_copy.columns]
                            prev_df_copy.columns = prev_cols_upper
                            name_col_p  = next((c for c in ["RESIDENT'S NAME","RESIDENTS NAME","NAME"] if c in prev_cols_upper), None)
                            street_col_p= next((c for c in ["STREET"] if c in prev_cols_upper), None)
                            occ_col_p   = next((c for c in ["OCCUPANCY","OCCUPANCY TYPE"] if c in prev_cols_upper), None)
                            rec_col_p   = next((c for c in ["RECEIPT NO.","RECEIPT NO","RECEIPT_NO"] if c in prev_cols_upper), None)
                            if name_col_p:
                                names_p = prev_df_copy[name_col_p].dropna().astype(str)
                                names_p = names_p[~names_p.str.upper().isin(['NAN','',name_col_p,"RESIDENT'S NAME",'NAME'])]
                                bfwd_c = prev_df_copy[rec_col_p].astype(str).str.upper().str.strip().isin(['BFWD','B/FWD']).sum() if rec_col_p else 0
                                pay_c  = len(names_p) - int(bfwd_c)
                                ep1,ep2,ep3,ep4 = st.columns(4)
                                with ep1: st.metric("👤 Residents",    int(len(names_p.unique())))
                                with ep2: st.metric("🏘️ Streets",      int(len(prev_df_copy[street_col_p].dropna().unique())) if street_col_p else 0)
                                with ep3: st.metric("📋 Balance rows", int(bfwd_c))
                                with ep4: st.metric("💰 Payment rows", int(pay_c))
                                if len(names_p.unique()) > 0:
                                    st.markdown("**👤 Residents found in file:**")
                                    st.caption(", ".join(sorted(names_p.unique().tolist())[:40]))
                                if street_col_p:
                                    streets_p = sorted(prev_df_copy[street_col_p].dropna().astype(str).str.upper().unique().tolist())
                                    st.markdown("**🏘️ Streets found in file:**")
                                    st.caption(", ".join(streets_p))
                                if occ_col_p:
                                    occ_p = prev_df_copy[occ_col_p].astype(str).str.upper().value_counts().to_dict()
                                    st.markdown("**👥 Occupancy types:**")
                                    st.caption(str({k:v for k,v in occ_p.items() if k not in ('NAN','')})  )
                            else:
                                st.warning("⚠️ Name column not detected. Make sure your file has a column called RESIDENT'S NAME.")

                        if st.button("🚀 Extract & Save to Database", use_container_width=True, type="primary"):
                            with st.spinner("Extracting data from file and saving each record to the database..."):
                                res = estate.process_excel_template_upload(preview_df)
                            cr1,cr2 = st.columns(2)
                            with cr1:
                                if res['success']>0:
                                    st.success(f"✅ **{res['success']}** records extracted from file and saved to database!")
                                if res['skipped']>0:
                                    st.info(f"⏭ **{res['skipped']}** rows had no payment amounts — skipped")
                            with cr2:
                                if res['warnings']:
                                    with st.expander(f"⚠️ {len(res['warnings'])} warning(s)"):
                                        for w in res['warnings']: st.warning(w)
                                if res['errors']:
                                    with st.expander(f"❌ {len(res['errors'])} error(s)"):
                                        for e in res['errors']: st.error(e)
                            if res['success']>0:
                                st.markdown("---")
                                st.markdown("**✅ Data extracted from Excel and saved into database tables:**")
                                r1,r2,r3 = st.columns(3)
                                with r1:
                                    st.metric("🏘️ Streets created",    res.get('streets_created',0))
                                    st.metric("🏠 Properties created", res.get('properties_created',0))
                                with r2:
                                    st.metric("👤 Residents created",  res.get('residents_created',0))
                                    st.metric("💰 Payments saved",     res.get('payments_created',0))
                                with r3:
                                    st.metric("📊 Balances (BFWD)",    res.get('balances_created',0))
                                    st.metric("⏭ Rows skipped",        res.get('skipped',0))
                                st.caption("Streets → streets table | Properties → properties table | Residents → residents table | Payments → payments table | BFWD rows → balances table")
                                st.balloons(); st.rerun()
                except Exception as e:
                    st.error(f"Error reading file: {e}"); st.error(traceback.format_exc())

        with col_info:
            st.markdown("""
            **Service Charge Amounts:**
            | Code | Name | Freq | Amount |
            |------|------|------|--------|
            | SVC | Service Charge | Annual | ₦240,000 |
            | DEV | Development Levy | Annual | ₦50,000 |
            | PARTY | Year End Party | Annual | ₦0 |
            | INFRA | Infrastructure | One Off | ₦2,000,000 |
            | LEGAL | Legal Fee | One Off | ₦50,000 |
            | TRANSF | Transformer | One Off | ₦50,000 |
            | LIGHT | Light Connection | One Off | ₦100,000 |

            **Notes:**
            - Landlords → ₦0 annual charges
            - Streets/properties/residents auto-created
            - Duplicate receipts safely skipped
            """)

    # ════════════════════════════════════════════════════════════
    # SETTINGS
    # ════════════════════════════════════════════════════════════
    elif selected == "⚙️ Settings":
        st.markdown('<div class="main-header"><h1>⚙️ Settings</h1></div>', unsafe_allow_html=True)
        tab1,tab2,tab3,tab4,tab5=st.tabs(["💳 Payment Types","🏠 House Types","💰 Payment Methods","🏦 Bank Accounts","🗑️ Clear Tables"])

        with tab1:
            st.subheader("Configure Service Charge Amounts")
            existing_pt=estate.get_payment_types()
            if existing_pt:
                for pt in existing_pt:
                    with st.expander(f"{pt['code']} — {pt['name']}"):
                        with st.form(f"pt_form_{pt['id']}"):
                            c1,c2=st.columns(2)
                            with c1:
                                new_name=st.text_input("Name",pt['name'])
                                new_desc=st.text_area("Description",safe_str(pt.get('description')))
                            with c2:
                                new_freq=st.selectbox("Frequency",['Annual','One Off'],
                                    index=0 if pt.get('charge_frequency')=='Annual' else 1)
                                new_amt=st.number_input("Default Amount (₦)",value=float(pt.get('default_amount',0)),min_value=0.0,step=1000.0)
                            if st.form_submit_button("💾 Save"):
                                ok,msg=estate.upsert_payment_type(pt['code'],new_name,new_desc,new_freq,new_amt)
                                if ok: st.success(msg); st.rerun()
                                else: st.error(msg)
            else: st.info("No payment types yet. Add one below or upload data.")
            st.markdown("---")
            st.subheader("Add New Payment Type")
            with st.form("add_pt"):
                c1,c2=st.columns(2)
                with c1:
                    new_code=st.text_input("Code (e.g. SVC)")
                    new_name_pt=st.text_input("Name")
                with c2:
                    new_freq_pt=st.selectbox("Frequency",['Annual','One Off'])
                    new_amt_pt=st.number_input("Amount (₦)",min_value=0.0,step=1000.0)
                new_desc_pt=st.text_input("Description")
                if st.form_submit_button("➕ Add"):
                    if new_code and new_name_pt:
                        ok,msg=estate.upsert_payment_type(new_code.upper(),new_name_pt,new_desc_pt,new_freq_pt,new_amt_pt)
                        if ok: st.success(msg); st.rerun()
                        else: st.error(msg)

        with tab2:
            st.subheader("🏠 House Types")
            st.caption("These are the property/house types extracted from the HOUSE TYPE column in your Excel file. No price — just the type name.")
            existing_ht = estate.get_property_types()
            if existing_ht:
                for ht_row in existing_ht:
                    c1,c2 = st.columns([4,1])
                    with c1: st.write(f"🏠 **{ht_row['name']}**")
                    with c2:
                        if st.button("🗑️", key=f"del_ht_{ht_row['id']}",
                                     help=f"Delete {ht_row['name']}"):
                            try:
                                cur_ht = estate.conn.cursor()
                                cur_ht.execute("DELETE FROM property_types WHERE id=%s", (ht_row['id'],))
                                estate.conn.commit(); cur_ht.close()
                                st.success(f"Deleted {ht_row['name']}"); st.rerun()
                            except Exception as e:
                                st.warning(f"Cannot delete — likely in use by a property. ({e})")
            else:
                st.info("No house types yet. They are created automatically when you upload your Excel file.")
            st.markdown("---")
            st.markdown("**Add a custom house type:**")
            with st.form("add_ht"):
                ht_name = st.text_input("House Type Name", placeholder="e.g. Bungalow, Duplex, Flat")
                if st.form_submit_button("➕ Add House Type"):
                    if ht_name.strip():
                        try:
                            cur_ht = estate.conn.cursor()
                            cur_ht.execute("INSERT IGNORE INTO property_types(name) VALUES(%s)", (ht_name.strip().title(),))
                            estate.conn.commit(); cur_ht.close()
                            st.success(f"✅ '{ht_name.strip().title()}' added"); st.rerun()
                        except Exception as e: st.error(str(e))
                    else: st.error("Please enter a name")

        with tab3:
            st.subheader("Payment Methods")
            for pm in estate.get_payment_methods():
                st.write(f"**{pm['name']}** — {pm.get('description','')}")
            st.markdown("---")
            with st.form("add_pm"):
                c1,c2=st.columns(2)
                with c1: pm_name=st.text_input("Method Name")
                with c2: pm_desc=st.text_input("Description")
                if st.form_submit_button("➕ Add Method"):
                    if pm_name:
                        ok,msg=estate.upsert_payment_method(pm_name,pm_desc)
                        if ok: st.success(msg); st.rerun()
                        else: st.error(msg)

        with tab4:
            st.subheader("Bank Accounts")
            for ba in estate.get_bank_accounts():
                c1,c2=st.columns([4,1])
                with c1: st.write(f"**{ba['bank_name']}** | {ba['account_name']} | `{ba['account_number']}` | {ba.get('branch','')}")
                with c2:
                    if st.button("🚫 Remove",key=f"del_ba_{ba['id']}"):
                        ok,msg=estate.delete_bank_account(ba['id'])
                        if ok: st.success(msg); st.rerun()
                        else: st.error(msg)
            st.markdown("---")
            with st.form("add_ba"):
                c1,c2=st.columns(2)
                with c1:
                    ba_bank=st.text_input("Bank Name")
                    ba_acct_name=st.text_input("Account Name")
                with c2:
                    ba_acct_no=st.text_input("Account Number")
                    ba_branch=st.text_input("Branch")
                if st.form_submit_button("➕ Add Account"):
                    if ba_bank and ba_acct_no:
                        ok,msg=estate.upsert_bank_account(ba_bank,ba_acct_name,ba_acct_no,ba_branch)
                        if ok: st.success(msg); st.rerun()
                        else: st.error(msg)

        with tab5:
            st.subheader("⚠️ Clear Database Tables")
            st.warning("**Danger Zone!** This permanently deletes ALL data from all tables. Cannot be undone.")
            confirm=st.text_input("Type **CLEAR ALL** to confirm",placeholder="CLEAR ALL")
            if st.button("🗑️ Clear All Tables",type="primary"):
                if confirm.strip()=="CLEAR ALL":
                    ok,msg=estate.clear_all_tables()
                    if ok: st.success(f"✅ {msg}"); st.rerun()
                    else: st.error(msg)
                else: st.error("Please type CLEAR ALL exactly to confirm")

    # Footer
    st.markdown("---")
    st.markdown("""<div style="text-align:center;color:#666;padding:10px">
        🏠 <strong>Sunshine Estate Management System</strong> v1.0 | Built with ❤️
    </div>""", unsafe_allow_html=True)

if __name__ == "__main__":
    main()